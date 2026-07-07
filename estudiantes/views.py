# estudiantes/views.py
import io
import bcrypt
from datetime import date, datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from .models import Estudiante, Acudiente, EstudianteAcudiente
from colegios.models import Colegio


# ============================================================
# HELPER — verificar sesión
# ============================================================
def _sesion_activa(request):
    return bool(request.session.get('usuario_cedula'))


def _rol_en(request, *roles):
    rol = request.session.get('usuario_rol')
    if rol == 'ADMIN':
        return True
    return rol in roles


# ============================================================
# MAPA LOCALIDAD → ZONA → RUTA AUTOMÁTICA
# ============================================================
LOCALIDAD_ZONA = {
    'USAQUEN':        'NORTE',
    'CHAPINERO':      'NORTE',
    'SUBA':           'NORTE',
    'BARRIOS_UNIDOS': 'CENTRO',
    'TEUSAQUILLO':    'CENTRO',
    'SANTA_FE':       'CENTRO',
    'CANDELARIA':     'CENTRO',
    'MARTIRES':       'CENTRO',
    'ANTONIO_NARINO': 'CENTRO',
    'PUENTE_ARANDA':  'OCCIDENTE',
    'KENNEDY':        'OCCIDENTE',
    'FONTIBON':       'OCCIDENTE',
    'ENGATIVA':       'OCCIDENTE',
    'SAN_CRISTOBAL':  'ORIENTE',
    'RAFAEL_URIBE':   'ORIENTE',
    'USME':           'SUR',
    'TUNJUELITO':     'SUR',
    'BOSA':           'SUR',
    'CIUDAD_BOLIVAR': 'SUR',
    'SUMAPAZ':        'SUR',
}


def _asignar_ruta_por_localidad(localidad):
    from rutas.models import Ruta
    if not localidad:
        return None
    zona = LOCALIDAD_ZONA.get(localidad)
    if not zona:
        return None
    ruta = Ruta.objects.filter(zona=zona, activo=True).first()
    return ruta.codigo if ruta else None


# ============================================================
# HELPER — Sincronizar parada del estudiante
# ============================================================
def _sincronizar_parada(estudiante):
    from rutas.models import Parada

    ref_tag = f'EST-{estudiante.documento}'
    parada_actual = Parada.objects.filter(referencia__contains=ref_tag).first()

    def _quitar_tag(parada):
        if not parada:
            return
        tags_restantes = [
            t.strip() for t in (parada.referencia or '').split(',')
            if t.strip() and ref_tag not in t
        ]
        if tags_restantes:
            parada.referencia = ', '.join(tags_restantes)
            parada.save(update_fields=['referencia'])
        else:
            parada.activo = False
            parada.referencia = ''
            parada.save(update_fields=['activo', 'referencia'])

    if not estudiante.activo or not estudiante.direccion or not estudiante.codigo_ruta_id:
        _quitar_tag(parada_actual)
        return None

    nombre_hijo = f'{estudiante.nombre} {estudiante.apellido}'.strip()

    parada_hermano = (
        Parada.objects
        .filter(ruta_id=estudiante.codigo_ruta_id, direccion__iexact=estudiante.direccion)
        .exclude(pk=getattr(parada_actual, 'pk', None))
        .first()
    )

    if parada_hermano:
        if parada_actual and parada_actual.pk != parada_hermano.pk:
            _quitar_tag(parada_actual)
        tags = [t.strip() for t in (parada_hermano.referencia or '').split(',') if t.strip()]
        if ref_tag not in tags:
            tags.append(ref_tag)
        parada_hermano.referencia = ', '.join(tags)
        parada_hermano.activo = True
        if nombre_hijo and nombre_hijo not in parada_hermano.nombre:
            parada_hermano.nombre = f'{parada_hermano.nombre} / {nombre_hijo}'
        parada_hermano.save()
        return parada_hermano

    if (
        parada_actual
        and parada_actual.direccion == estudiante.direccion
        and parada_actual.ruta_id == estudiante.codigo_ruta_id
    ):
        parada_actual.nombre = f'Casa de {nombre_hijo}'
        parada_actual.activo = True
        parada_actual.save()
        return parada_actual

    if parada_actual:
        _quitar_tag(parada_actual)

    ultimo_orden = (
        Parada.objects.filter(ruta_id=estudiante.codigo_ruta_id)
        .order_by('-orden')
        .values_list('orden', flat=True)
        .first()
    )

    return Parada.objects.create(
        ruta_id    = estudiante.codigo_ruta_id,
        orden      = (ultimo_orden or 0) + 1,
        nombre     = f'Casa de {nombre_hijo}',
        direccion  = estudiante.direccion,
        referencia = ref_tag,
        latitud    = None,
        longitud   = None,
        activo     = True,
    )


# ============================================================
# HELPER — Enviar correo de bienvenida con credenciales
# ============================================================
def _generar_y_enviar_invitacion(request, acudiente, nombre_hijo):
    """
    Envía correo de bienvenida al padre/acudiente con sus credenciales.
    El usuario ya queda activo desde el momento del registro.
    Devuelve True si el correo se envió con éxito, False si falló.
    """
    url_login = 'https://web-production-cadfa.up.railway.app/login/'

    try:
        from django.core.mail import EmailMultiAlternatives
        from django.conf import settings as cfg

        asunto = '🚌 Bienvenido/a a SafeRoute — Tus credenciales de acceso'

        texto_plano = (
            f'Hola {acudiente.nombre},\n\n'
            f'El colegio ha registrado a {nombre_hijo} en SafeRoute.\n'
            f'Tu cuenta ya está activa. Tus credenciales son:\n\n'
            f'  Usuario:    {acudiente.documento}\n'
            f'  Contraseña: {acudiente.documento}\n'
            f'  Acceso:     {url_login}\n\n'
            f'Por seguridad, cambia tu contraseña después del primer ingreso.\n\n'
            f'Con tu cuenta podrás:\n'
            f'  • Ver la ubicación del bus en tiempo real\n'
            f'  • Recibir notificaciones de recogida y entrega\n'
            f'  • Reportar ausencias y notificar al conductor\n'
            f'  • Consultar y gestionar pagos del servicio\n\n'
            f'— Equipo SafeRoute'
        )

        iniciales = ''.join([p[0].upper() for p in nombre_hijo.split()[:2]])

        html = f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f6f8fc;font-family:Arial,sans-serif;">
  <div style="max-width:520px;margin:32px auto;">

    <div style="background:#1e293b;border-radius:12px 12px 0 0;padding:28px;text-align:center;">
      <div style="font-size:36px;">🚌</div>
      <div style="font-size:20px;font-weight:800;color:white;margin-top:8px;">SafeRoute</div>
      <div style="font-size:11px;color:#94a3b8;text-transform:uppercase;letter-spacing:0.05em;">
        Sistema de Gestión de Transporte Escolar
      </div>
      <div style="margin-top:20px;font-size:18px;font-weight:700;color:white;">
        ¡Bienvenido/a a SafeRoute!
      </div>
      <div style="font-size:12px;color:#94a3b8;margin-top:4px;">
        El colegio te ha registrado como acudiente
      </div>
    </div>

    <div style="background:white;padding:28px;">

      <div style="font-size:16px;font-weight:700;color:#1f2937;margin-bottom:8px;">
        Hola, <span style="color:#f59e0b;">{acudiente.nombre}</span> 👋
      </div>
      <p style="font-size:13px;color:#4b5563;line-height:1.6;margin:0 0 20px;">
        El colegio ha registrado a <strong>{nombre_hijo}</strong> en el sistema de
        transporte escolar. Tu cuenta ya está activa y puedes ingresar ahora mismo.
      </p>

      <div style="background:#f8fafc;border:1.5px solid #e2e8f0;border-radius:10px;padding:14px;margin-bottom:20px;">
        <div style="font-size:9px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:0.07em;margin-bottom:10px;">
          Estudiante vinculado
        </div>
        <div style="display:flex;align-items:center;gap:12px;">
          <div style="width:38px;height:38px;border-radius:50%;background:#f59e0b;display:flex;align-items:center;justify-content:center;font-size:13px;font-weight:700;color:white;flex-shrink:0;">
            {iniciales}
          </div>
          <div>
            <div style="font-size:13px;font-weight:600;color:#1f2937;">{nombre_hijo}</div>
            <div style="font-size:11px;color:#6b7280;">Estudiante registrado en SafeRoute</div>
          </div>
          <div style="margin-left:auto;background:#d1fae5;color:#065f46;font-size:10px;font-weight:600;padding:2px 8px;border-radius:20px;">Activo</div>
        </div>
      </div>

      <div style="background:#fef3c7;border:1.5px solid #fcd34d;border-radius:10px;padding:16px;margin-bottom:20px;">
        <div style="font-size:10px;font-weight:700;color:#92400e;text-transform:uppercase;letter-spacing:0.06em;margin-bottom:12px;">
          Tus credenciales de acceso
        </div>
        <div style="display:flex;justify-content:space-between;align-items:center;padding:6px 0;border-bottom:1px solid #fde68a;">
          <span style="font-size:12px;color:#92400e;">🔒 Usuario</span>
          <span style="font-size:13px;font-weight:700;color:#78350f;font-family:monospace;">{acudiente.documento}</span>
        </div>
        <div style="display:flex;justify-content:space-between;align-items:center;padding:6px 0;">
          <span style="font-size:12px;color:#92400e;">🔑 Contraseña</span>
          <span style="font-size:13px;font-weight:700;color:#78350f;font-family:monospace;">{acudiente.documento}</span>
        </div>
      </div>

      <div style="text-align:center;margin:24px 0 20px;">
        <a href="{url_login}"
           style="display:inline-block;background:#f59e0b;color:white;font-weight:700;font-size:14px;padding:14px 36px;border-radius:8px;text-decoration:none;">
          🚀 &nbsp; Ingresar a SafeRoute
        </a>
      </div>

      <div style="font-size:10px;font-weight:700;color:#374151;text-transform:uppercase;letter-spacing:0.06em;margin-bottom:10px;">
        Con tu cuenta podrás:
      </div>
      <div style="font-size:12px;color:#4b5563;padding:4px 0;"><span style="color:#f59e0b;font-weight:700;">✓</span> &nbsp;Ver la ubicación del bus en tiempo real</div>
      <div style="font-size:12px;color:#4b5563;padding:4px 0;"><span style="color:#f59e0b;font-weight:700;">✓</span> &nbsp;Recibir notificaciones de recogida y entrega</div>
      <div style="font-size:12px;color:#4b5563;padding:4px 0;"><span style="color:#f59e0b;font-weight:700;">✓</span> &nbsp;Reportar ausencias y notificar al conductor</div>
      <div style="font-size:12px;color:#4b5563;padding:4px 0;"><span style="color:#f59e0b;font-weight:700;">✓</span> &nbsp;Consultar y gestionar pagos del servicio</div>

      <div style="background:#fef3c7;border:1px solid #fcd34d;border-radius:8px;padding:12px 14px;font-size:11px;color:#92400e;line-height:1.6;margin-top:20px;">
        <strong>⚠️ Recomendación de seguridad:</strong> cambia tu contraseña después de tu primer ingreso.
      </div>

    </div>

    <div style="background:#f8fafc;border-top:1px solid #e2e8f0;padding:14px 28px;text-align:center;font-size:11px;color:#6b7280;line-height:1.6;">
      Si no tienes un hijo matriculado en este colegio, puedes ignorar este mensaje.
    </div>

    <div style="background:#1e293b;border-radius:0 0 12px 12px;padding:20px;text-align:center;">
      <div style="font-size:14px;font-weight:800;color:white;">🚌 SafeRoute</div>
      <div style="font-size:10px;color:#64748b;margin-top:8px;line-height:1.7;">
        Sistema de Gestión de Transporte Escolar<br>
        Bogotá D.C. · Colombia · © 2026 SafeRoute<br>
        Correo enviado a <span style="color:#94a3b8;">{acudiente.email}</span>
      </div>
    </div>

  </div>
</body>
</html>"""

        correo = EmailMultiAlternatives(
            subject=asunto,
            body=texto_plano,
            from_email=getattr(cfg, 'DEFAULT_FROM_EMAIL', 'noreply@saferoute.co'),
            to=[acudiente.email],
        )
        correo.attach_alternative(html, 'text/html')
        correo.send(fail_silently=False)
        return True

    except Exception as e:
        print(f'⚠️  No se pudo enviar correo a {acudiente.email}: {e}')
        return False


# ============================================================
# HELPER — Procesar un acudiente (principal o secundario)
# ============================================================
def _procesar_acudiente(request, prefijo, nombre_hijo, es_principal):
    from usuarios.models import Usuario

    tipo_doc   = request.POST.get(f'{prefijo}_tipo_documento', '').strip()
    documento  = request.POST.get(f'{prefijo}_documento', '').strip()
    nombre     = request.POST.get(f'{prefijo}_nombre', '').strip()
    email      = request.POST.get(f'{prefijo}_email', '').strip().lower()
    telefono   = request.POST.get(f'{prefijo}_telefono', '').strip()
    tipo_notif = request.POST.get(f'{prefijo}_notificaciones', 'TODAS')

    if not documento or not email:
        return None, None

    acudiente = Acudiente.objects.filter(documento=documento).first()

    if acudiente:
        if telefono and acudiente.telefono != telefono:
            acudiente.telefono = telefono
            acudiente.save()

        if not acudiente.usuario or not acudiente.usuario.activo:
            if not acudiente.usuario:
                try:
                    hash_pass = bcrypt.hashpw(acudiente.documento.encode(), bcrypt.gensalt()).decode()
                    usuario_nuevo = Usuario.objects.create(
                        cedula         = acudiente.documento,
                        tipo_documento = acudiente.tipo_documento,
                        user_name      = acudiente.documento,
                        password       = hash_pass,
                        nombre         = acudiente.nombre,
                        email          = acudiente.email,
                        telefono       = acudiente.telefono or None,
                        rol            = 'PADRE',
                        activo         = True,
                    )
                    acudiente.usuario = usuario_nuevo
                    acudiente.save()
                except Exception as e:
                    print(f'Error creando usuario para acudiente existente: {e}')
            else:
                acudiente.usuario.activo = True
                acudiente.usuario.save(update_fields=['activo'])

            envio_ok = _generar_y_enviar_invitacion(request, acudiente, nombre_hijo)
            if envio_ok:
                messages.info(
                    request,
                    f'El acudiente {acudiente.nombre} ya existe. '
                    f'Se reenvió el correo con credenciales a {acudiente.email}.'
                )
            else:
                messages.warning(
                    request,
                    f'El acudiente {acudiente.nombre} ya existe, pero no se pudo '
                    f'enviar el correo a {acudiente.email}. '
                    f'Sus credenciales son: usuario y contraseña = {acudiente.documento}.'
                )
    else:
        acudiente = Acudiente.objects.create(
            documento      = documento,
            tipo_documento = tipo_doc or 'CEDULA_CIUDADANIA',
            nombre         = nombre or f'Acudiente de {nombre_hijo}',
            email          = email,
            telefono       = telefono or None,
        )

        hash_pass = bcrypt.hashpw(documento.encode(), bcrypt.gensalt()).decode()

        usuario_padre = Usuario.objects.create(
            cedula         = documento,
            tipo_documento = tipo_doc or 'CEDULA_CIUDADANIA',
            user_name      = documento,
            password       = hash_pass,
            nombre         = nombre or f'Acudiente de {nombre_hijo}',
            email          = email,
            telefono       = telefono or None,
            rol            = 'PADRE',
            activo         = True,
        )
        acudiente.usuario = usuario_padre
        acudiente.save()

        envio_ok = _generar_y_enviar_invitacion(request, acudiente, nombre_hijo)
        if envio_ok:
            messages.success(
                request,
                f'✅ Acudiente registrado. Correo con credenciales enviado a {email}.'
            )
        else:
            messages.warning(
                request,
                f'Acudiente registrado, pero no se pudo enviar el correo a {email}. '
                f'Sus credenciales son: usuario y contraseña = {documento}.'
            )

    return acudiente, tipo_notif


# ============================================================
# LISTAR ESTUDIANTES
# ============================================================
def lista_estudiantes(request):
    if not _sesion_activa(request):
        return redirect('login')

    try:
        from rutas.models import Ruta

        rol = request.session.get('usuario_rol')
        es_personal_operativo = rol in ('CONDUCTOR', 'MONITORA')

        nombre      = request.GET.get('nombre', '')
        grado       = request.GET.get('grado', '')
        institucion = request.GET.get('institucion', '')
        activo      = request.GET.get('activo', 'true')
        ruta        = request.GET.get('ruta', '')

        estudiantes = Estudiante.objects.all().order_by('nombre')

        if es_personal_operativo:
            estudiantes = estudiantes.filter(activo=True)
            activo = 'true'
        else:
            if activo == 'true':
                estudiantes = estudiantes.filter(activo=True)
            elif activo == 'false':
                estudiantes = estudiantes.filter(activo=False)

        if nombre:
            estudiantes = estudiantes.filter(nombre__icontains=nombre)
        if grado:
            estudiantes = estudiantes.filter(grado__icontains=grado)
        if institucion:
            estudiantes = estudiantes.filter(institucion__icontains=institucion)
        if ruta:
            estudiantes = estudiantes.filter(codigo_ruta=ruta)

        rutas = Ruta.objects.filter(activo=True)

        context = {
            'estudiantes':          estudiantes,
            'rutas':                rutas,
            'colegios':             Colegio.objects.filter(activo=True).order_by('nombre_institucion'),
            'total_estudiantes':    Estudiante.objects.count(),
            'estudiantes_activos':  Estudiante.objects.filter(activo=True).count(),
            'estudiantes_sin_ruta': Estudiante.objects.filter(codigo_ruta__isnull=True).count(),
            'usuario_nombre':       request.session.get('usuario_nombre'),
            'usuario_rol':          rol,
            'puede_ver_historial':  not es_personal_operativo,
            'nombre':      nombre,
            'grado':       grado,
            'institucion': institucion,
            'activo':      activo,
            'ruta':        ruta,
        }
    except Exception as e:
        context = {
            'estudiantes':    [],
            'rutas':          [],
            'usuario_nombre': request.session.get('usuario_nombre'),
            'usuario_rol':    request.session.get('usuario_rol'),
            'error':          f'Error al cargar estudiantes: {str(e)}',
        }

    return render(request, 'estudiantes/lista_estudiantes.html', context)


# ============================================================
# CREAR ESTUDIANTE
# ============================================================
def nuevo_estudiante(request):
    if not _sesion_activa(request):
        return redirect('login')

    if not _rol_en(request, 'COLEGIO', 'ADMIN'):
        messages.error(request, 'No tienes permisos para registrar estudiantes.')
        return redirect('estudiantes')

    if request.method == 'POST':
        try:
            documento      = request.POST.get('documento', '').strip()
            tipo_documento = request.POST.get('tipo_documento', '').strip()
            nombre         = request.POST.get('nombre', '').strip()
            apellido       = request.POST.get('apellido', '').strip()
            fecha_nac      = request.POST.get('fecha_nacimiento', '').strip()
            grado          = request.POST.get('grado', '').strip()
            institucion    = request.POST.get('institucion', '').strip()
            tipo_sangre    = request.POST.get('tipo_sangre', '').strip()
            direccion      = request.POST.get('direccion', '').strip()
            contacto_nom   = request.POST.get('contacto_emergencia_nombre', '').strip()
            contacto_tel   = request.POST.get('contacto_emergencia_telefono', '').strip()
            enfermedades   = request.POST.get('enfermedades', '').strip()

            localidad   = request.POST.get('localidad', '').strip() or None
            # La ruta SIEMPRE se asigna automáticamente según la localidad
            codigo_ruta = _asignar_ruta_por_localidad(localidad) if localidad else None

            if Estudiante.objects.filter(documento=documento).exists():
                messages.error(request, 'Ya existe un estudiante con ese documento.')
                return redirect('estudiantes')

            nombre_hijo = f'{nombre} {apellido}'
            acudiente1, notif1 = _procesar_acudiente(
                request, 'acudiente1', nombre_hijo, es_principal=True
            )
            acudiente2, notif2 = _procesar_acudiente(
                request, 'acudiente2', nombre_hijo, es_principal=False
            )

            cedula_padre_id = None
            if acudiente1 and acudiente1.usuario:
                cedula_padre_id = acudiente1.usuario.cedula

            estudiante = Estudiante.objects.create(
                documento                    = documento,
                tipo_documento               = tipo_documento,
                nombre                       = nombre,
                apellido                     = apellido,
                fecha_nacimiento             = fecha_nac,
                grado                        = grado or None,
                institucion                  = institucion or None,
                tipo_sangre                  = tipo_sangre or None,
                direccion                    = direccion or None,
                localidad                    = localidad or None,
                contacto_emergencia_nombre   = contacto_nom or None,
                contacto_emergencia_telefono = contacto_tel or None,
                codigo_ruta_id               = codigo_ruta,
                cedula_padre_id              = cedula_padre_id,
                enfermedades                 = enfermedades or None,
                activo                       = True,
            )

            _sincronizar_parada(estudiante)

            if acudiente1:
                EstudianteAcudiente.objects.create(
                    estudiante          = estudiante,
                    acudiente           = acudiente1,
                    es_principal        = True,
                    tipo_notificaciones = notif1 or 'TODAS',
                )

            if acudiente2:
                EstudianteAcudiente.objects.create(
                    estudiante          = estudiante,
                    acudiente           = acudiente2,
                    es_principal        = False,
                    tipo_notificaciones = notif2 or 'CRITICAS',
                )

            messages.success(
                request,
                f'✅ Estudiante {nombre} {apellido} registrado exitosamente.'
            )

        except Exception as e:
            messages.error(request, f'Error al registrar estudiante: {str(e)}')

    return redirect('estudiantes')


# ============================================================
# EDITAR ESTUDIANTE
# ============================================================
def estudiante_editar(request, documento):
    if not _sesion_activa(request):
        return redirect('login')

    try:
        estudiante = Estudiante.objects.get(documento=documento)
    except Estudiante.DoesNotExist:
        messages.error(request, 'Estudiante no encontrado.')
        return redirect('estudiantes')

    if request.method == 'POST':
        try:
            estudiante.nombre   = request.POST.get('nombre', estudiante.nombre).strip()
            estudiante.apellido = request.POST.get('apellido', estudiante.apellido).strip()
            fecha_nac = request.POST.get('fecha_nacimiento', '').strip()
            if fecha_nac:
                estudiante.fecha_nacimiento = fecha_nac
            estudiante.grado                        = request.POST.get('grado', '').strip() or None
            estudiante.institucion                  = request.POST.get('institucion', '').strip() or None
            estudiante.tipo_sangre                  = request.POST.get('tipo_sangre', '').strip() or None
            estudiante.direccion                    = request.POST.get('direccion', '').strip() or None
            estudiante.localidad                    = request.POST.get('localidad', '').strip() or None
            estudiante.contacto_emergencia_nombre   = request.POST.get('contacto_emergencia_nombre', '').strip() or None
            estudiante.contacto_emergencia_telefono = request.POST.get('contacto_emergencia_telefono', '').strip() or None
            estudiante.enfermedades                 = request.POST.get('enfermedades', '').strip() or None
            # La ruta se recalcula automáticamente cada vez, según la localidad actual
            estudiante.codigo_ruta_id = _asignar_ruta_por_localidad(estudiante.localidad) if estudiante.localidad else None
            activo = request.POST.get('activo', 'true')
            estudiante.activo = (activo == 'true')
            estudiante.save()

            _sincronizar_parada(estudiante)

            messages.success(
                request,
                f'✅ Estudiante {estudiante.nombre} {estudiante.apellido} actualizado.'
            )
            return redirect('estudiantes')
        except Exception as e:
            messages.error(request, f'Error al editar: {str(e)}')

    try:
        from rutas.models import Ruta
        rutas = Ruta.objects.filter(activo=True)
    except Exception:
        rutas = []

    try:
        colegios = Colegio.objects.filter(activo=True).order_by('nombre_institucion')
    except Exception:
        colegios = []

    context = {
        'es_nuevo':       False,
        'estudiante':     estudiante,
        'rutas':          rutas,
        'colegios':       colegios,
        'fecha_actual':   date.today(),
        'usuario_nombre': request.session.get('usuario_nombre'),
        'usuario_rol':    request.session.get('usuario_rol'),
    }
    return render(request, 'estudiantes/nuevo_estudiante.html', context)


# ============================================================
# ELIMINAR ESTUDIANTE (SOFT DELETE)
# ============================================================
def estudiante_eliminar(request, documento):
    if not _sesion_activa(request):
        return redirect('login')

    if request.method != 'POST':
        return redirect('estudiantes')

    try:
        estudiante = Estudiante.objects.get(documento=documento)
        nombre     = f'{estudiante.nombre} {estudiante.apellido}'
        estudiante.activo = False
        estudiante.save(update_fields=['activo'])
        _sincronizar_parada(estudiante)
        messages.success(request, f'Estudiante {nombre} desactivado.')
    except Estudiante.DoesNotExist:
        messages.error(request, 'Estudiante no encontrado.')
    except Exception as e:
        messages.error(request, f'Error al desactivar: {str(e)}')

    return redirect('estudiantes')


# ============================================================
# DESCARGAR PDF — Estudiantes
# ============================================================
def estudiantes_pdf(request):
    if not _sesion_activa(request):
        return redirect('login')

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib           import colors
        from reportlab.lib.units     import cm
        from reportlab.platypus      import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle

        nombre      = request.GET.get('nombre', '')
        grado       = request.GET.get('grado', '')
        institucion = request.GET.get('institucion', '')
        activo      = request.GET.get('activo', '')
        ruta        = request.GET.get('ruta', '')

        estudiantes = Estudiante.objects.all().order_by('nombre')
        if nombre:
            estudiantes = estudiantes.filter(nombre__icontains=nombre)
        if grado:
            estudiantes = estudiantes.filter(grado__icontains=grado)
        if institucion:
            estudiantes = estudiantes.filter(institucion__icontains=institucion)
        if activo == 'true':
            estudiantes = estudiantes.filter(activo=True)
        elif activo == 'false':
            estudiantes = estudiantes.filter(activo=False)
        if ruta:
            estudiantes = estudiantes.filter(codigo_ruta=ruta)

        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            leftMargin=1*cm, rightMargin=1*cm,
            topMargin=1.5*cm, bottomMargin=1*cm
        )
        styles = getSampleStyleSheet()
        story  = []

        titulo_style = ParagraphStyle(
            't', parent=styles['Title'], fontSize=14,
            textColor=colors.HexColor('#1e293b')
        )
        story.append(Paragraph('SafeRoute — Reporte de Estudiantes', titulo_style))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(
            f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Total: {estudiantes.count()} registros',
            styles['Normal']
        ))
        story.append(Spacer(1, 0.5*cm))

        cabeceras = ['Documento', 'Nombre', 'Apellido', 'Grado', 'Institución', 'RH', 'Estado']
        filas = [
            [e.documento, e.nombre, e.apellido, e.grado or '—',
             e.institucion or '—', e.tipo_sangre or '—',
             'Activo' if e.activo else 'Inactivo']
            for e in estudiantes
        ]
        if not filas:
            filas = [['Sin datos'] * len(cabeceras)]

        page_w = landscape(A4)[0] - 2*cm
        col_w  = [page_w / len(cabeceras)] * len(cabeceras)
        tabla  = Table([cabeceras] + filas, colWidths=col_w, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#1e293b')),
            ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',       (0, 0), (-1, 0),  9),
            ('ALIGN',          (0, 0), (-1, 0),  'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('FONTSIZE',       (0, 1), (-1, -1), 8),
            ('GRID',           (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',     (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING',  (0, 0), (-1, -1), 4),
        ]))
        story.append(tabla)
        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="estudiantes_{date.today()}.pdf"'
        return response

    except Exception as e:
        messages.error(request, f'Error al generar PDF: {str(e)}')
        return redirect('estudiantes')


# ============================================================
# DESCARGAR EXCEL — Estudiantes
# ============================================================
def estudiantes_excel(request):
    if not _sesion_activa(request):
        return redirect('login')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        nombre      = request.GET.get('nombre', '')
        grado       = request.GET.get('grado', '')
        institucion = request.GET.get('institucion', '')
        activo      = request.GET.get('activo', '')
        ruta        = request.GET.get('ruta', '')

        estudiantes = Estudiante.objects.all().order_by('nombre')
        if nombre:
            estudiantes = estudiantes.filter(nombre__icontains=nombre)
        if grado:
            estudiantes = estudiantes.filter(grado__icontains=grado)
        if institucion:
            estudiantes = estudiantes.filter(institucion__icontains=institucion)
        if activo == 'true':
            estudiantes = estudiantes.filter(activo=True)
        elif activo == 'false':
            estudiantes = estudiantes.filter(activo=False)
        if ruta:
            estudiantes = estudiantes.filter(codigo_ruta=ruta)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Estudiantes'

        header_fill = PatternFill('solid', fgColor='1E293B')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        thin        = Side(style='thin', color='E2E8F0')
        borde       = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill    = PatternFill('solid', fgColor='F8FAFC')

        ws.merge_cells('A1:H1')
        ws['A1']           = f'SafeRoute — Estudiantes | {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A1'].font      = Font(bold=True, size=12, color='1E293B')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:H2')
        ws['A2']           = f'Total: {estudiantes.count()} registros'
        ws['A2'].alignment = Alignment(horizontal='center')

        cabeceras = ['Documento', 'Tipo Doc', 'Nombre', 'Apellido', 'Grado', 'Institución', 'Tipo Sangre', 'Estado']
        for col, cab in enumerate(cabeceras, 1):
            cell           = ws.cell(row=4, column=col, value=cab)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = borde
        ws.row_dimensions[4].height = 22

        for row_idx, e in enumerate(estudiantes, 5):
            fila = [e.documento, e.tipo_documento, e.nombre, e.apellido,
                    e.grado or '—', e.institucion or '—',
                    e.tipo_sangre or '—', 'Activo' if e.activo else 'Inactivo']
            fill = alt_fill if row_idx % 2 == 0 else None
            for col, valor in enumerate(fila, 1):
                cell           = ws.cell(row=row_idx, column=col, value=valor)
                cell.border    = borde
                cell.alignment = Alignment(vertical='center')
                if fill:
                    cell.fill = fill

        anchos = [15, 20, 20, 20, 15, 25, 12, 10]
        letras = ['A','B','C','D','E','F','G','H']
        for letra, ancho in zip(letras, anchos):
            ws.column_dimensions[letra].width = ancho

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="estudiantes_{date.today()}.xlsx"'
        return response

    except Exception as e:
        messages.error(request, f'Error al generar Excel: {str(e)}')
        return redirect('estudiantes')