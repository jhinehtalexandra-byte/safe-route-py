# pagos/views.py
import io
import json
import uuid
import hashlib
from datetime import date, datetime
from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from .models import Pago


def _filtrar_pagos(pagos_qs, request):
    estado      = request.GET.get('estado',      '').strip()
    metodo      = request.GET.get('metodo',      '').strip()
    mes         = request.GET.get('mes',         '').strip()
    anio        = request.GET.get('anio',        '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    if estado:      pagos_qs = pagos_qs.filter(estado=estado)
    if metodo:      pagos_qs = pagos_qs.filter(metodo_pago=metodo)
    if mes:         pagos_qs = pagos_qs.filter(mes__icontains=mes)
    if anio:        pagos_qs = pagos_qs.filter(anio=anio)
    if fecha_desde: pagos_qs = pagos_qs.filter(fecha_pago__gte=fecha_desde)
    if fecha_hasta: pagos_qs = pagos_qs.filter(fecha_pago__lte=fecha_hasta)
    return pagos_qs


def _pagos_base_admin():
    return Pago.objects.all().order_by('-fecha_registro')


def _pagos_base_padre(cedula):
    return Pago.objects.filter(cedula_padre=cedula).order_by('-fecha_registro')


# ── GESTIÓN DE PAGOS (Admin/Colegio) ──────────────────────────────────────────
def lista_pagos(request):
    if not request.session.get('usuario_cedula'):
        return redirect('login')
    rol = request.session.get('usuario_rol')
    try:
        pagos_qs = _filtrar_pagos(_pagos_base_admin(), request)
        base     = _pagos_base_admin()
        context = {
            'pagos':            pagos_qs,
            'pagos_pendientes': base.filter(estado='PENDIENTE').count(),
            'pagos_realizados': base.filter(estado='PAGADO').count(),
            'pagos_vencidos':   base.filter(estado='VENCIDO').count(),
            'total_pagado':     sum(p.monto for p in base.filter(estado='PAGADO')),
            'usuario_nombre':   request.session.get('usuario_nombre'),
            'usuario_rol':      rol,
            'estado':           request.GET.get('estado', ''),
            'metodo':           request.GET.get('metodo', ''),
            'mes':              request.GET.get('mes', ''),
            'fecha_desde':      request.GET.get('fecha_desde', ''),
            'fecha_hasta':      request.GET.get('fecha_hasta', ''),
            'fecha_actual':     date.today(),
        }
    except Exception as e:
        context = {
            'pagos': [], 'usuario_nombre': request.session.get('usuario_nombre'),
            'usuario_rol': rol, 'error': f'Error: {str(e)}', 'fecha_actual': date.today(),
        }
    return render(request, 'pagos/gestion_pagos.html', context)


# ── PAGOS EN LÍNEA (Padre) ────────────────────────────────────────────────────
def pagos_en_linea(request):
    if not request.session.get('usuario_cedula'):
        return redirect('login')
    rol    = request.session.get('usuario_rol')
    cedula = request.session.get('usuario_cedula')
    if rol not in ('PADRE', 'ADMIN'):
        messages.error(request, 'No tienes permisos para esta sección.')
        return redirect('login')

    try:
        base                   = _pagos_base_padre(cedula) if rol == 'PADRE' else _pagos_base_admin()
        pagos_filtrados        = _filtrar_pagos(base, request)
        pagos_pendientes_lista = base.filter(estado__in=('PENDIENTE', 'VENCIDO')).order_by('fecha_vencimiento')
        hijos = []
        try:
            from estudiantes.models import Estudiante
            qs_hijos = Estudiante.objects.filter(activo=True) if rol == 'ADMIN' else \
                       Estudiante.objects.filter(cedula_padre=cedula, activo=True)
            hijos = list(qs_hijos.select_related('codigo_ruta'))
        except Exception:
            pass
        context = {
            'pagos_pendientes_lista': pagos_pendientes_lista,
            'pagos':                  pagos_filtrados,
            'pagos_pendientes':       pagos_pendientes_lista.count(),
            'pagos_realizados':       base.filter(estado='PAGADO').count(),
            'pagos_vencidos':         base.filter(estado='VENCIDO').count(),
            'total_pagado':           sum(p.monto for p in base.filter(estado='PAGADO')),
            'usuario_nombre':         request.session.get('usuario_nombre'),
            'usuario_rol':            rol,
            'fecha_actual':           date.today(),
            'hijos':                  hijos,
            'estado':                 request.GET.get('estado', ''),
            'anio':                   request.GET.get('anio', ''),
            'wompi_public_key':       settings.WOMPI_PUBLIC_KEY,
        }
    except Exception as e:
        context = {
            'pagos_pendientes_lista': [], 'pagos': [],
            'pagos_pendientes': 0, 'pagos_realizados': 0,
            'pagos_vencidos': 0, 'total_pagado': 0,
            'error': f'Error: {str(e)}',
            'usuario_nombre': request.session.get('usuario_nombre'),
            'usuario_rol': rol, 'fecha_actual': date.today(), 'hijos': [],
            'wompi_public_key': settings.WOMPI_PUBLIC_KEY,
        }
    return render(request, 'pagos/pagos_en_linea.html', context)


# ── CREAR PAGO INDIVIDUAL (Admin/Colegio) ─────────────────────────────────────
def crear_pago(request):
    if not request.session.get('usuario_cedula'):
        return redirect('login')
    rol = request.session.get('usuario_rol')
    if rol not in ('ADMIN', 'COLEGIO'):
        return redirect('login')

    if request.method == 'POST':
        try:
            from estudiantes.models import Estudiante

            documento_est = request.POST.get('documento_estudiante', '').strip()
            monto         = request.POST.get('monto', '').strip()
            mes           = request.POST.get('mes', '').strip()
            anio          = request.POST.get('anio', '').strip()
            concepto      = request.POST.get('concepto', '').strip()
            fecha_venc    = request.POST.get('fecha_vencimiento', '').strip() or None

            if not all([documento_est, monto, mes, anio]):
                messages.error(request, 'Documento, monto, mes y año son obligatorios.')
                return redirect('crear_pago')

            estudiante = Estudiante.objects.get(documento=documento_est)

            if not estudiante.cedula_padre:
                messages.error(
                    request,
                    f'El estudiante {estudiante.nombre} no tiene acudiente/padre asignado.'
                )
                return redirect('crear_pago')

            codigo = f"PAG-{documento_est[:6]}-{mes[:3].upper()}{anio}-{uuid.uuid4().hex[:4].upper()}"

            Pago.objects.create(
                codigo               = codigo,
                monto                = float(monto),
                mes                  = mes,
                anio                 = int(anio),
                concepto             = concepto or f'Transporte escolar {mes} {anio}',
                estado               = 'PENDIENTE',
                fecha_vencimiento    = fecha_venc,
                documento_estudiante = estudiante,
                cedula_padre         = estudiante.cedula_padre,
            )
            messages.success(request, f'✅ Pago {codigo} creado para {estudiante.nombre}.')
            return redirect('pagos')

        except Estudiante.DoesNotExist:
            messages.error(request, 'Estudiante no encontrado.')
        except Exception as e:
            messages.error(request, f'Error al crear pago: {str(e)}')
        return redirect('crear_pago')

    # GET — cargar formulario
    try:
        from estudiantes.models import Estudiante
        estudiantes = Estudiante.objects.filter(activo=True).order_by('nombre')
    except Exception:
        estudiantes = []

    meses = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
    ]

    return render(request, 'pagos/crear_pago.html', {
        'estudiantes':    estudiantes,
        'meses':          meses,
        'anio_actual':    date.today().year,
        'usuario_nombre': request.session.get('usuario_nombre'),
        'usuario_rol':    rol,
        'fecha_actual':   date.today(),
    })


# ── CREAR PAGOS MASIVOS (Admin/Colegio) ───────────────────────────────────────
def crear_pagos_masivos(request):
    """Genera pagos pendientes para TODOS los estudiantes activos de un mes/año."""
    if not request.session.get('usuario_cedula'):
        return redirect('login')
    if request.session.get('usuario_rol') not in ('ADMIN', 'COLEGIO'):
        return redirect('login')

    if request.method == 'POST':
        try:
            from estudiantes.models import Estudiante

            mes        = request.POST.get('mes', '').strip()
            anio       = request.POST.get('anio', '').strip()
            monto      = request.POST.get('monto', '').strip()
            fecha_venc = request.POST.get('fecha_vencimiento', '').strip() or None

            if not all([mes, anio, monto]):
                messages.error(request, 'Mes, año y monto son obligatorios.')
                return redirect('pagos')

            estudiantes = Estudiante.objects.filter(activo=True).exclude(cedula_padre=None)
            creados  = 0
            omitidos = 0

            for est in estudiantes:
                # Evitar duplicados del mismo mes/año/estudiante
                if Pago.objects.filter(documento_estudiante=est, mes=mes, anio=int(anio)).exists():
                    omitidos += 1
                    continue

                codigo = f"PAG-{est.documento[:6]}-{mes[:3].upper()}{anio}-{uuid.uuid4().hex[:4].upper()}"
                Pago.objects.create(
                    codigo               = codigo,
                    monto                = float(monto),
                    mes                  = mes,
                    anio                 = int(anio),
                    concepto             = f'Transporte escolar {mes} {anio}',
                    estado               = 'PENDIENTE',
                    fecha_vencimiento    = fecha_venc,
                    documento_estudiante = est,
                    cedula_padre         = est.cedula_padre,
                )
                creados += 1

            messages.success(
                request,
                f'✅ {creados} pagos creados para {mes} {anio}. '
                f'{omitidos} omitidos (ya existían).'
            )

        except Exception as e:
            messages.error(request, f'Error al crear pagos masivos: {str(e)}')

    return redirect('pagos')


# ── PROCESAR PAGO (Admin desde gestion_pagos) ─────────────────────────────────
def procesar_pago(request):
    if not request.session.get('usuario_cedula'):
        return redirect('login')
    if request.method == 'POST':
        try:
            codigo_pago = request.POST.get('codigo_pago', '').strip()
            metodo_pago = request.POST.get('metodo_pago', 'EFECTIVO').strip()
            if codigo_pago:
                pago = Pago.objects.get(codigo=codigo_pago)
                pago.estado     = 'PAGADO'
                pago.metodo_pago = metodo_pago
                pago.fecha_pago  = date.today()
                pago.save()
                messages.success(request, f'Pago {codigo_pago} procesado.')
            else:
                messages.error(request, 'Código de pago no encontrado.')
        except Pago.DoesNotExist:
            messages.error(request, 'Pago no encontrado.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    return redirect('pagos')


# ════════════════════════════════════════════════════════════════════════════
#  WOMPI — Integración de pasarela de pagos
# ════════════════════════════════════════════════════════════════════════════

def datos_widget_wompi(request, codigo_pago):
    """
    Devuelve en JSON la referencia única y la firma de integridad
    necesarias para abrir el Widget de Wompi desde el frontend.
    """
    if not request.session.get('usuario_cedula'):
        return JsonResponse({'ok': False, 'error': 'Sesión no válida.'}, status=401)

    rol    = request.session.get('usuario_rol')
    cedula = request.session.get('usuario_cedula')

    try:
        filtro = {'codigo': codigo_pago, 'estado__in': ('PENDIENTE', 'VENCIDO')}
        if rol == 'PADRE':
            filtro['cedula_padre'] = cedula
        pago = Pago.objects.get(**filtro)
    except Pago.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Pago no encontrado o ya fue procesado.'}, status=404)

    referencia = f"PAGO-{pago.codigo}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    pago.referencia_wompi = referencia
    pago.save(update_fields=['referencia_wompi'])

    monto_centavos = int(pago.monto * 100)
    moneda = "COP"

    cadena_firma     = f"{referencia}{monto_centavos}{moneda}{settings.WOMPI_INTEGRITY_SECRET}"
    firma_integridad = hashlib.sha256(cadena_firma.encode('utf-8')).hexdigest()

    return JsonResponse({
        'ok':               True,
        'referencia':       referencia,
        'monto_centavos':   monto_centavos,
        'moneda':           moneda,
        'firma_integridad': firma_integridad,
        'wompi_public_key': settings.WOMPI_PUBLIC_KEY,
        'redirect_url':     request.build_absolute_uri('/pagos/wompi/confirmacion/'),
    })


@csrf_exempt
@require_POST
def webhook_wompi(request):
    """Wompi llama esta URL cuando una transacción cambia de estado."""
    try:
        payload = json.loads(request.body)
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    evento      = payload.get('event')
    data        = payload.get('data', {})
    transaccion = data.get('transaction', {})

    if evento != 'transaction.updated':
        return JsonResponse({'ok': True})

    # ── Validar firma del webhook ──
    timestamp          = payload.get('timestamp')
    checksum_recibido  = payload.get('signature', {}).get('checksum')
    propiedades        = payload.get('signature', {}).get('properties', [])

    valores_concatenados = ""
    for prop in propiedades:
        partes = prop.split('.')
        valor  = transaccion
        for parte in partes[1:]:
            valor = valor.get(parte, '') if isinstance(valor, dict) else ''
        valores_concatenados += str(valor)

    cadena_verificacion = f"{valores_concatenados}{timestamp}{settings.WOMPI_EVENTS_SECRET}"
    checksum_calculado  = hashlib.sha256(cadena_verificacion.encode('utf-8')).hexdigest()

    if checksum_calculado != checksum_recibido:
        return JsonResponse({'error': 'Firma inválida'}, status=403)

    # ── Actualizar pago ──
    referencia     = transaccion.get('reference', '')
    estado_wompi   = transaccion.get('status', '')
    transaccion_id = transaccion.get('id', '')
    metodo         = transaccion.get('payment_method_type', '')

    try:
        pago = Pago.objects.get(referencia_wompi=referencia)
    except Pago.DoesNotExist:
        return JsonResponse({'error': 'Pago no encontrado'}, status=404)

    pago.estado_wompi         = estado_wompi
    pago.transaccion_id_wompi = transaccion_id

    mapa_metodo = {
        'CARD':                   'TARJETA_CREDITO',
        'PSE':                    'PSE',
        'NEQUI':                  'NEQUI',
        'BANCOLOMBIA_TRANSFER':   'TRANSFERENCIA',
    }

    if estado_wompi == 'APPROVED':
        pago.estado      = 'PAGADO'
        pago.fecha_pago  = date.today()
        pago.metodo_pago = mapa_metodo.get(metodo, pago.metodo_pago)
        pago.comprobante = transaccion_id
    elif estado_wompi in ('DECLINED', 'ERROR'):
        pago.estado = 'PENDIENTE'

    pago.save()
    return JsonResponse({'ok': True})


def confirmacion_pago_wompi(request):
    """Página informativa a la que Wompi redirige al cerrar el widget."""
    if not request.session.get('usuario_cedula'):
        return redirect('login')
    referencia = request.GET.get('id', '')
    pago       = Pago.objects.filter(referencia_wompi=referencia).first()
    return render(request, 'pagos/confirmacion_wompi.html', {
        'pago': pago,
        'usuario_rol': request.session.get('usuario_rol'),
    })


# ── PDF ───────────────────────────────────────────────────────────────────────
def pagos_pdf(request):
    if not request.session.get('usuario_cedula'):
        return redirect('login')
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        rol    = request.session.get('usuario_rol')
        cedula = request.session.get('usuario_cedula')
        pagos  = _filtrar_pagos(
            _pagos_base_padre(cedula) if rol == 'PADRE' else _pagos_base_admin(),
            request
        )
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            leftMargin=1*cm, rightMargin=1*cm,
            topMargin=1.5*cm, bottomMargin=1*cm,
        )
        styles = getSampleStyleSheet()
        story  = []
        story.append(Paragraph(
            'SafeRoute — Reporte de Pagos',
            ParagraphStyle('t', parent=styles['Title'], fontSize=14,
                           textColor=colors.HexColor('#1e293b'))
        ))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(
            f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Total: {pagos.count()} registros',
            styles['Normal']
        ))
        story.append(Spacer(1, 0.5*cm))

        cabeceras = ['Código', 'Estudiante', 'Monto', 'Mes / Año', 'Fecha Pago', 'Método', 'Estado']
        filas = [
            [
                p.codigo,
                str(p.documento_estudiante_id),
                f'${p.monto}',
                f'{p.mes or ""} {p.anio or ""}'.strip(),
                p.fecha_pago.strftime('%d/%m/%Y') if p.fecha_pago else '—',
                p.metodo_pago or '—',
                p.estado,
            ]
            for p in pagos
        ]
        if not filas:
            filas = [['Sin datos'] + [''] * (len(cabeceras) - 1)]

        pw    = landscape(A4)[0] - 2*cm
        tabla = Table([cabeceras] + filas, colWidths=[pw / len(cabeceras)] * len(cabeceras), repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1e293b')),
            ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, 0),  9),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('FONTSIZE',      (0, 1), (-1, -1), 8),
            ('GRID',          (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(tabla)
        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="pagos_{date.today()}.pdf"'
        return response

    except Exception as e:
        messages.error(request, f'Error al generar PDF: {str(e)}')
        return redirect('pagos_en_linea' if request.session.get('usuario_rol') == 'PADRE' else 'pagos')


# ── EXCEL ─────────────────────────────────────────────────────────────────────
def pagos_excel(request):
    if not request.session.get('usuario_cedula'):
        return redirect('login')
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        rol    = request.session.get('usuario_rol')
        cedula = request.session.get('usuario_cedula')
        pagos  = _filtrar_pagos(
            _pagos_base_padre(cedula) if rol == 'PADRE' else _pagos_base_admin(),
            request
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Pagos'

        hf   = PatternFill('solid', fgColor='1E293B')
        hfnt = Font(color='FFFFFF', bold=True, size=10)
        thin = Side(style='thin', color='E2E8F0')
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt  = PatternFill('solid', fgColor='F8FAFC')

        ws.merge_cells('A1:G1')
        ws['A1']           = f'SafeRoute — Pagos | {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A1'].font      = Font(bold=True, size=12, color='1E293B')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:G2')
        ws['A2']           = f'Total: {pagos.count()} registros'
        ws['A2'].alignment = Alignment(horizontal='center')

        cabeceras = ['Código', 'Estudiante', 'Monto', 'Mes', 'Año', 'Fecha Pago', 'Estado']
        for col, cab in enumerate(cabeceras, 1):
            c           = ws.cell(row=4, column=col, value=cab)
            c.fill      = hf
            c.font      = hfnt
            c.border    = brd
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[4].height = 22

        for ri, p in enumerate(pagos, 5):
            fila = [
                p.codigo,
                str(p.documento_estudiante_id),
                float(p.monto),
                p.mes or '—',
                p.anio or '—',
                p.fecha_pago.strftime('%d/%m/%Y') if p.fecha_pago else '—',
                p.estado,
            ]
            fill = alt if ri % 2 == 0 else None
            for col, val in enumerate(fila, 1):
                c           = ws.cell(row=ri, column=col, value=val)
                c.border    = brd
                c.alignment = Alignment(vertical='center')
                if fill:
                    c.fill = fill

        for col in ws.columns:
            ancho = 10
            for cell in col:
                try:
                    if cell.value:
                        ancho = max(ancho, len(str(cell.value)))
                except Exception:
                    pass
            try:
                ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 40)
            except Exception:
                pass

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="pagos_{date.today()}.xlsx"'
        return response

    except Exception as e:
        messages.error(request, f'Error al generar Excel: {str(e)}')
        return redirect('pagos_en_linea' if request.session.get('usuario_rol') == 'PADRE' else 'pagos')