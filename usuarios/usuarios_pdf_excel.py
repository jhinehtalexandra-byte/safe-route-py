"""
AGREGA estas dos funciones al final de tu usuarios/views.py existente,
y actualiza la función 'usuarios' con el contexto extra para las estadísticas.
"""

# ============================================
# ACTUALIZA la función usuarios() — agrega estas líneas al context:
# ============================================
#
#   context = {
#       'usuarios':          lista,
#       'usuario_nombre':    request.session.get('usuario_nombre'),
#       'busqueda':          busqueda,
#       'rol':               rol,
#       'estado':            estado,
#       # ── NUEVAS LÍNEAS ──
#       'total_usuarios':    Usuario.objects.count(),
#       'usuarios_activos':  Usuario.objects.filter(activo=True).count(),
#       'total_conductores': Usuario.objects.filter(rol='CONDUCTOR').count(),
#   }
#
# ============================================


# ============================================
# FUNCIÓN AUXILIAR — filtrar usuarios (para PDF y Excel)
# ============================================
def _filtrar_usuarios(request):
    from .models import Usuario
    busqueda = request.GET.get('busqueda', '')
    rol      = request.GET.get('rol', '')
    estado   = request.GET.get('estado', '')

    qs = Usuario.objects.all().order_by('nombre')
    if busqueda:
        qs = qs.filter(nombre__icontains=busqueda)
    if rol:
        qs = qs.filter(rol=rol)
    if estado == 'activo':
        qs = qs.filter(activo=True)
    elif estado == 'inactivo':
        qs = qs.filter(activo=False)

    return list(qs)


# ============================================
# DESCARGAR PDF — Usuarios
# ============================================
def usuarios_pdf(request):
    import io
    from datetime import date, datetime
    from django.http import HttpResponse
    from django.contrib import messages
    from django.shortcuts import redirect

    if not request.session.get('usuario_cedula'):
        return redirect('login')
    if request.session.get('usuario_rol') != 'ADMIN':
        return redirect('login')

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib           import colors
        from reportlab.lib.units     import cm
        from reportlab.platypus      import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle

        datos  = _filtrar_usuarios(request)
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=1*cm, rightMargin=1*cm,
            topMargin=1.5*cm, bottomMargin=1*cm
        )
        styles = getSampleStyleSheet()
        story  = []

        titulo_style = ParagraphStyle('titulo', parent=styles['Title'], fontSize=16,
                                      textColor=colors.HexColor('#1e293b'))
        story.append(Paragraph('RutaEscolar — Reporte de Usuarios', titulo_style))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(
            f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Total: {len(datos)} registros',
            styles['Normal']
        ))
        story.append(Spacer(1, 0.5*cm))

        cabeceras = ['Cédula', 'Nombre', 'Usuario', 'Email', 'Teléfono', 'Rol', 'Estado']
        filas = []
        for d in datos:
            filas.append([
                str(d.cedula),
                d.nombre,
                d.user_name,
                d.email,
                d.telefono or '—',
                d.rol,
                'Activo' if d.activo else 'Inactivo',
            ])

        if not filas:
            filas = [['Sin datos para los filtros seleccionados'] + [''] * (len(cabeceras) - 1)]

        tabla_data = [cabeceras] + filas
        page_w     = landscape(A4)[0] - 2*cm
        col_w      = [page_w / len(cabeceras)] * len(cabeceras)

        tabla = Table(tabla_data, colWidths=col_w, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1e293b')),
            ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, 0),  9),
            ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('FONTSIZE',      (0, 1), (-1, -1), 8),
            ('GRID',          (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))

        story.append(tabla)
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph('RutaEscolar © 2025 | Sistema de Gestión de Transporte Escolar', styles['Normal']))
        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_usuarios_{date.today()}.pdf"'
        return response

    except Exception as e:
        messages.error(request, f'Error al generar PDF: {str(e)}')
        return redirect('usuarios')


# ============================================
# DESCARGAR EXCEL — Usuarios
# ============================================
def usuarios_excel(request):
    import io
    from datetime import date, datetime
    from django.http import HttpResponse
    from django.contrib import messages
    from django.shortcuts import redirect

    if not request.session.get('usuario_cedula'):
        return redirect('login')
    if request.session.get('usuario_rol') != 'ADMIN':
        return redirect('login')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        datos = _filtrar_usuarios(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Usuarios'

        header_fill  = PatternFill('solid', fgColor='1e293b')
        header_font  = Font(color='FFFFFF', bold=True, size=10)
        header_align = Alignment(horizontal='center', vertical='center')
        thin         = Side(style='thin', color='e2e8f0')
        thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill     = PatternFill('solid', fgColor='f8fafc')

        ws.merge_cells('A1:G1')
        ws['A1'] = f'RutaEscolar — Reporte de Usuarios | {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A1'].font      = Font(bold=True, size=13, color='1e293b')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        ws.merge_cells('A2:G2')
        ws['A2'] = f'Total registros: {len(datos)}'
        ws['A2'].alignment = Alignment(horizontal='center')

        cabeceras = ['Cédula', 'Nombre', 'Usuario', 'Email', 'Teléfono', 'Rol', 'Estado']
        for col, cab in enumerate(cabeceras, 1):
            cell = ws.cell(row=4, column=col, value=cab)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_align
            cell.border    = thin_border
        ws.row_dimensions[4].height = 22

        for row_idx, d in enumerate(datos, 5):
            fila = [d.cedula, d.nombre, d.user_name, d.email,
                    d.telefono or '—', d.rol, 'Activo' if d.activo else 'Inactivo']
            fill = alt_fill if row_idx % 2 == 0 else None
            for col, valor in enumerate(fila, 1):
                cell = ws.cell(row=row_idx, column=col, value=valor)
                cell.border    = thin_border
                cell.alignment = Alignment(vertical='center')
                if fill:
                    cell.fill = fill

        for col in ws.columns:
            max_len    = max((len(str(cell.value or '')) for cell in col), default=10)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_usuarios_{date.today()}.xlsx"'
        return response

    except Exception as e:
        messages.error(request, f'Error al generar Excel: {str(e)}')
        return redirect('usuarios')