# usuarios/views.py
import io
import re          # CORRECCIÓN: movido al nivel de módulo (estaba dentro de una función)
import bcrypt
import secrets
from datetime import datetime, date, timedelta
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.core.mail import send_mail
from django.conf import settings
from .models import Usuario


# ============================================================
# HELPERS
# ============================================================
def _sesion_activa(request):
    return bool(request.session.get('usuario_cedula'))


def _es_admin(request):
    return request.session.get('usuario_rol') == 'ADMIN'


def _rol_en(request, *roles):
    """El ADMIN siempre tiene acceso total."""
    rol = request.session.get('usuario_rol')
    if rol == 'ADMIN':
        return True
    return rol in roles


def _es_admin_o_colegio(request):
    return request.session.get('usuario_rol') in ('ADMIN', 'COLEGIO')


def _redirigir_dashboard(request):
    """Redirige al dashboard correcto según el rol."""
    DASHBOARDS = {
        'ADMIN':     'dashboard_admin',
        'COLEGIO':   'dashboard_colegio',
        'CONDUCTOR': 'dashboard_conductor',
        'MONITORA':  'dashboard_monitora',
        'PADRE':     'dashboard_padre',
    }
    rol = request.session.get('usuario_rol', '')
    return redirect(DASHBOARDS.get(rol, 'home'))


# ============================================================
# HOME
# ============================================================
def home(request):
    return render(request, 'home.html')


# ============================================================
# LOGIN
# ============================================================
def login_view(request):
    if _sesion_activa(request):
        return _redirigir_dashboard(request)

    if request.method == 'POST':
        user_name = request.POST.get('username', '').strip()
        password  = request.POST.get('password', '').strip()

        if not user_name or not password:
            return render(request, 'login.html', {
                'error': 'Por favor ingresa usuario y contraseña.'
            })

        try:
            usuario = Usuario.objects.get(user_name=user_name, activo=True)
            password_correcta = bcrypt.checkpw(
                password.encode('utf-8'),
                usuario.password.encode('utf-8')
            )

            if password_correcta:
                request.session['usuario_cedula']   = usuario.cedula
                request.session['usuario_nombre']   = usuario.nombre
                request.session['usuario_rol']      = usuario.rol
                request.session['lista_confirmada'] = False
                request.session.cycle_key()
                return _redirigir_dashboard(request)
            else:
                return render(request, 'login.html', {
                    'error': 'Usuario o contraseña incorrectos.'
                })

        except Usuario.DoesNotExist:
            return render(request, 'login.html', {
                'error': 'Usuario o contraseña incorrectos.'
            })
        except Exception as e:
            return render(request, 'login.html', {
                'error': f'Error inesperado: {str(e)}'
            })

    return render(request, 'login.html')


# ============================================================
# LOGOUT
# ============================================================
def logout_view(request):
    request.session.flush()
    messages.success(request, 'Sesión cerrada exitosamente.')
    return redirect('login')


# ============================================================
# CONTACTO
# ============================================================
def contacto(request):
    if request.method == 'POST':
        messages.success(request, '¡Mensaje enviado exitosamente!')
    return redirect('home')


# ============================================================
# RECUPERAR CONTRASEÑA — Paso 1: solicitar enlace
# ============================================================
def recuperar_password(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()

        if not email:
            return render(request, 'recuperar_password.html', {
                'error': 'Debes ingresar un correo electrónico.',
                'paso': 1
            })

        usuario = Usuario.objects.filter(email=email, activo=True).first()

        if usuario:
            token  = secrets.token_urlsafe(32)
            expira = (datetime.now() + timedelta(minutes=30)).isoformat()

            request.session[f'reset_token_{token}'] = {
                'cedula': usuario.cedula,
                'expira': expira,
            }

            dominio = request.build_absolute_uri('/')[:-1]
            enlace  = f"{dominio}/restablecer-password/{token}/"

            try:
                send_mail(
                    subject='🔐 Recuperar contraseña — RutaEscolar',
                    message=(
                        f'Hola {usuario.nombre},\n\n'
                        f'Recibimos una solicitud para restablecer tu contraseña.\n\n'
                        f'Haz clic en el siguiente enlace (válido por 30 minutos):\n'
                        f'{enlace}\n\n'
                        f'Si no solicitaste esto, ignora este correo.\n\n'
                        f'— Equipo RutaEscolar'
                    ),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    fail_silently=False,
                )
            except Exception as e:
                return render(request, 'recuperar_password.html', {
                    'error': f'No se pudo enviar el correo. Verifica la configuración: {str(e)}',
                    'paso': 1,
                    'email_previo': email,
                })

        # Respuesta genérica por seguridad (no revelar si el correo existe o no)
        return render(request, 'recuperar_password.html', {
            'exito':         True,
            'email_enviado': email,
            'paso':          2,
        })

    return render(request, 'recuperar_password.html', {'paso': 1})


# ============================================================
# RECUPERAR CONTRASEÑA — Paso 2: formulario nueva contraseña
# ============================================================
def restablecer_password(request, token):
    datos_token = request.session.get(f'reset_token_{token}')

    if not datos_token:
        messages.error(request, 'El enlace no es válido o ya fue usado.')
        return redirect('recuperar_password')

    expira = datetime.fromisoformat(datos_token['expira'])
    if datetime.now() > expira:
        del request.session[f'reset_token_{token}']
        messages.error(request, 'El enlace ha expirado. Solicita uno nuevo.')
        return redirect('recuperar_password')

    if request.method == 'POST':
        nueva     = request.POST.get('password', '').strip()
        confirmar = request.POST.get('confirm_password', '').strip()

        # CORRECCIÓN: `import re` movido al inicio del archivo;
        # aquí se usa directamente sin reimportarlo.
        if len(nueva) < 10:
            return render(request, 'restablecer_password.html', {
                'token': token,
                'error': 'La contraseña debe tener al menos 10 caracteres.',
            })
        if not re.search(r'[A-Z]', nueva):
            return render(request, 'restablecer_password.html', {
                'token': token,
                'error': 'La contraseña debe contener al menos una mayúscula.',
            })
        if not re.search(r'[a-z]', nueva):
            return render(request, 'restablecer_password.html', {
                'token': token,
                'error': 'La contraseña debe contener al menos una minúscula.',
            })
        if not re.search(r'[0-9]', nueva):
            return render(request, 'restablecer_password.html', {
                'token': token,
                'error': 'La contraseña debe contener al menos un número.',
            })
        if nueva != confirmar:
            return render(request, 'restablecer_password.html', {
                'token': token,
                'error': 'Las contraseñas no coinciden.',
            })

        try:
            usuario = Usuario.objects.get(cedula=datos_token['cedula'])
            usuario.password = bcrypt.hashpw(
                nueva.encode('utf-8'), bcrypt.gensalt()
            ).decode('utf-8')
            usuario.save()

            del request.session[f'reset_token_{token}']

            messages.success(request, '✅ Contraseña restablecida exitosamente. Ya puedes iniciar sesión.')
            return redirect('login')

        except Usuario.DoesNotExist:
            messages.error(request, 'Usuario no encontrado.')
            return redirect('recuperar_password')

    return render(request, 'restablecer_password.html', {'token': token})


# ============================================================
# TÉRMINOS Y CONDICIONES
# ============================================================
def terminos(request):
    return render(request, 'terminos.html')


# ============================================================
# POLÍTICA DE PRIVACIDAD
# ============================================================
def privacidad(request):
    return render(request, 'privacidad.html')


# ============================================================
# ERROR 404
# ============================================================
def error_404(request, exception):
    return render(request, '404.html', status=404)


# ============================================================
# DASHBOARD ADMIN
# ============================================================
def dashboard_admin(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _es_admin(request):
        return _redirigir_dashboard(request)

    try:
        from estudiantes.models import Estudiante
        from rutas.models import Ruta
        from pagos.models import Pago

        context = {
            'fecha_actual':      date.today(),
            'usuario_nombre':    request.session.get('usuario_nombre'),
            'usuario_rol':       request.session.get('usuario_rol'),
            'total_estudiantes': Estudiante.objects.filter(activo=True).count(),
            'total_rutas':       Ruta.objects.filter(activo=True).count(),
            'ingresos_mes':      Pago.objects.filter(estado='PAGADO').count(),
            'pagos_pendientes':  Pago.objects.filter(estado='PENDIENTE').count(),
            'proximas_rutas':    Ruta.objects.filter(activo=True)[:3],
        }
    except Exception as e:
        context = {
            'fecha_actual':   date.today(),
            'usuario_nombre': request.session.get('usuario_nombre'),
            'usuario_rol':    request.session.get('usuario_rol'),
            'error':          f'Error al cargar datos: {str(e)}',
        }

    return render(request, 'dashboard_admin.html', context)


# ============================================================
# DASHBOARD COLEGIO
# ============================================================
def dashboard_colegio(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'COLEGIO'):
        return _redirigir_dashboard(request)

    try:
        from estudiantes.models import Estudiante
        from rutas.models import Ruta
        from pagos.models import Pago

        context = {
            'usuario_nombre':    request.session.get('usuario_nombre'),
            'usuario_rol':       request.session.get('usuario_rol'),
            'fecha_actual':      date.today(),
            'total_estudiantes': Estudiante.objects.filter(activo=True).count(),
            'total_rutas':       Ruta.objects.filter(activo=True).count(),
            'total_conductores': Usuario.objects.filter(rol='CONDUCTOR', activo=True).count(),
            'pagos_pendientes':  Pago.objects.filter(estado='PENDIENTE').count(),
            'rutas_activas':     Ruta.objects.filter(activo=True)[:5],
            'novedades':         [],
            'notif_pendientes':  0,
        }
    except Exception as e:
        context = {
            'usuario_nombre': request.session.get('usuario_nombre'),
            'usuario_rol':    request.session.get('usuario_rol'),
            'fecha_actual':   date.today(),
            'error':          f'Error al cargar datos: {str(e)}',
        }

    return render(request, 'dashboard_colegio.html', context)


# ============================================================
# DASHBOARD CONDUCTOR
# ============================================================
def dashboard_conductor(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'CONDUCTOR'):
        return _redirigir_dashboard(request)

    try:
        from rutas.models import Ruta, Recorrido, ParadaRecorrido

        cedula    = request.session.get('usuario_cedula')
        conductor = Usuario.objects.get(cedula=cedula)

        ruta_actual = Ruta.objects.filter(
            conductor_cedula=conductor, activo=True
        ).first()

        recorrido_hoy = None
        paradas       = []
        paradas_completadas = 0

        if ruta_actual:
            recorrido_hoy = Recorrido.objects.filter(
                ruta  = ruta_actual,
                fecha = date.today(),
            ).exclude(estado='CANCELADO').order_by('-fecha_creacion').first()

            if not recorrido_hoy:
                recorrido_hoy = Recorrido.objects.create(
                    ruta      = ruta_actual,
                    conductor = conductor,
                    fecha     = date.today(),
                    estado    = 'PENDIENTE',
                    lista_confirmada = request.session.get('lista_confirmada', False),
                )

            if recorrido_hoy:
                paradas_qs = ParadaRecorrido.objects.filter(
                    recorrido=recorrido_hoy
                ).select_related('parada').order_by('parada__orden')
                paradas = list(paradas_qs)
                paradas_completadas = sum(1 for p in paradas if p.estado == 'COMPLETADA')

        total_estudiantes_ruta = 0
        if ruta_actual:
            from estudiantes.models import Estudiante
            total_estudiantes_ruta = Estudiante.objects.filter(
                codigo_ruta=ruta_actual, activo=True
            ).count()

        context = {
            'usuario_nombre':        request.session.get('usuario_nombre'),
            'usuario_rol':           request.session.get('usuario_rol'),
            'fecha_actual':          date.today(),
            'hora_actual':           datetime.now(),
            'ruta_actual':           ruta_actual,
            'recorrido_hoy':         recorrido_hoy,
            'paradas':               paradas,
            'paradas_completadas':   paradas_completadas,
            'total_rutas_hoy':       1 if ruta_actual else 0,
            'total_estudiantes_ruta': total_estudiantes_ruta,
            'total_paradas':         len(paradas),
            'alertas':               [],
            'alertas_count':         0,
        }

    except Exception as e:
        context = {
            'usuario_nombre':  request.session.get('usuario_nombre'),
            'usuario_rol':     request.session.get('usuario_rol'),
            'fecha_actual':    date.today(),
            'hora_actual':     datetime.now(),
            'ruta_actual':     None,
            'recorrido_hoy':   None,
            'paradas':         [],
            'paradas_completadas': 0,
            'error':           f'Error al cargar datos: {str(e)}',
        }

    return render(request, 'dashboard_conductor.html', context)


# ============================================================
# DASHBOARD MONITORA
# ============================================================
def dashboard_monitora(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'MONITORA'):
        return _redirigir_dashboard(request)

    try:
        from estudiantes.models import Estudiante

        estudiantes = Estudiante.objects.filter(activo=True).order_by('nombre')

        context = {
            'usuario_nombre':            request.session.get('usuario_nombre'),
            'usuario_rol':               request.session.get('usuario_rol'),
            'fecha_actual':              date.today(),
            'estudiantes':               estudiantes,
            'lista_confirmada':          request.session.get('lista_confirmada', False),
            'notif_pendientes':          0,
            'autorizaciones_pendientes': [],
        }
    except Exception as e:
        context = {
            'usuario_nombre':   request.session.get('usuario_nombre'),
            'usuario_rol':      request.session.get('usuario_rol'),
            'fecha_actual':     date.today(),
            'estudiantes':      [],
            'lista_confirmada': False,
            'error':            f'Error al cargar datos: {str(e)}',
        }

    return render(request, 'dashboard_monitora.html', context)


# ============================================================
# DASHBOARD PADRE
# ============================================================
def dashboard_padre(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'PADRE'):
        return _redirigir_dashboard(request)

    try:
        from estudiantes.models import Estudiante
        from pagos.models import Pago

        cedula = request.session.get('usuario_cedula')

        if _es_admin(request):
            hijos       = Estudiante.objects.filter(activo=True)[:5]
            pagos_lista = Pago.objects.filter(estado='PENDIENTE')[:5]
        else:
            hijos       = Estudiante.objects.filter(cedula_padre=cedula, activo=True)
            pagos_lista = Pago.objects.filter(cedula_padre=cedula, estado='PENDIENTE')

        # CORRECCIÓN: antes se hacía hijos.first().codigo_ruta sin comprobar
        # que existiera un primer elemento, lo que lanzaba AttributeError
        # cuando el padre no tenía hijos registrados.
        primer_hijo   = hijos.first()
        ruta_asignada = primer_hijo.codigo_ruta if primer_hijo else None

        context = {
            'usuario_nombre':   request.session.get('usuario_nombre'),
            'usuario_rol':      request.session.get('usuario_rol'),
            'fecha_actual':     date.today(),
            'total_hijos':      hijos.count(),
            'hijos':            hijos,
            'pagos_lista':      pagos_lista,
            'pagos_pendientes': pagos_lista.count(),
            'ruta_asignada':    ruta_asignada,
        }
    except Exception as e:
        context = {
            'usuario_nombre': request.session.get('usuario_nombre'),
            'usuario_rol':    request.session.get('usuario_rol'),
            'fecha_actual':   date.today(),
            'error':          f'Error al cargar datos: {str(e)}',
        }

    return render(request, 'dashboard_padre.html', context)


# ============================================================
# REGISTRARSE
# ============================================================
def registrarse(request):
    if request.method == 'POST':
        try:
            cedula         = request.POST.get('cedula', '').strip()
            tipo_documento = request.POST.get('tipo_documento', '').strip()
            user_name      = request.POST.get('userName', '').strip()
            password       = request.POST.get('password', '').strip()
            nombre         = request.POST.get('nombre', '').strip()
            email          = request.POST.get('email', '').strip().lower()
            telefono       = request.POST.get('telefono', '').strip()

            if not all([cedula, tipo_documento, user_name, password, nombre, email]):
                return render(request, 'registrarse.html', {
                    'error': 'Todos los campos obligatorios deben completarse.',
                    'form_data': request.POST,
                })

            if Usuario.objects.filter(user_name=user_name).exists():
                return render(request, 'registrarse.html', {
                    'error': 'El nombre de usuario ya existe.',
                    'form_data': request.POST,
                })
            if Usuario.objects.filter(email=email).exists():
                return render(request, 'registrarse.html', {
                    'error': 'El correo electrónico ya está registrado.',
                    'form_data': request.POST,
                })
            if Usuario.objects.filter(cedula=cedula).exists():
                return render(request, 'registrarse.html', {
                    'error': 'La cédula ya está registrada.',
                    'form_data': request.POST,
                })

            password_hash = bcrypt.hashpw(
                password.encode('utf-8'),
                bcrypt.gensalt()
            ).decode('utf-8')

            Usuario.objects.create(
                cedula         = cedula,
                tipo_documento = tipo_documento,
                user_name      = user_name,
                password       = password_hash,
                nombre         = nombre,
                email          = email,
                telefono       = telefono or None,
                rol            = 'PADRE',
                activo         = True,
            )
            messages.success(request, 'Registro exitoso. Ya puedes iniciar sesión.')
            return redirect('login')

        except Exception as e:
            return render(request, 'registrarse.html', {
                'error': f'Error al registrar: {str(e)}',
                'form_data': request.POST,
            })

    return render(request, 'registrarse.html')


# ============================================================
# PERFIL
# ============================================================
def perfil(request):
    if not _sesion_activa(request):
        return redirect('login')

    try:
        from estudiantes.models import Estudiante
        from pagos.models import Pago

        cedula  = request.session.get('usuario_cedula')
        usuario = Usuario.objects.get(cedula=cedula)

        if request.method == 'POST':
            accion = request.POST.get('accion')

            if accion == 'actualizar_perfil':
                usuario.nombre   = request.POST.get('nombre', usuario.nombre).strip()
                usuario.email    = request.POST.get('email', usuario.email).strip()
                usuario.telefono = request.POST.get('telefono', '').strip() or None
                usuario.save()
                request.session['usuario_nombre'] = usuario.nombre
                messages.success(request, 'Perfil actualizado exitosamente.')

            elif accion == 'cambiar_password':
                password_actual = request.POST.get('password_actual', '')
                password_nueva  = request.POST.get('password_nueva', '')

                if bcrypt.checkpw(
                    password_actual.encode('utf-8'),
                    usuario.password.encode('utf-8')
                ):
                    if len(password_nueva) < 10:
                        messages.error(request, 'La nueva contraseña debe tener al menos 10 caracteres.')
                    else:
                        usuario.password = bcrypt.hashpw(
                            password_nueva.encode('utf-8'),
                            bcrypt.gensalt()
                        ).decode('utf-8')
                        usuario.save()
                        messages.success(request, 'Contraseña cambiada exitosamente.')
                else:
                    messages.error(request, 'La contraseña actual es incorrecta.')

            return redirect('perfil')

        dias_activo = (
            (date.today() - usuario.fecha_creacion.date()).days
            if usuario.fecha_creacion else 0
        )

        context = {
            'usuario':           usuario,
            'usuario_nombre':    usuario.nombre,
            'usuario_rol':       usuario.rol,
            'fecha_actual':      date.today(),
            'total_estudiantes': Estudiante.objects.filter(cedula_padre=cedula).count(),
            'total_pagos':       Pago.objects.filter(cedula_padre=cedula).count(),
            'dias_activo':       dias_activo,
        }

    except Usuario.DoesNotExist:
        return redirect('login')
    except Exception as e:
        context = {
            'usuario_nombre': request.session.get('usuario_nombre'),
            'usuario_rol':    request.session.get('usuario_rol'),
            'error':          f'Error al cargar perfil: {str(e)}',
        }

    return render(request, 'perfil.html', context)


# ============================================================
# CRUD USUARIOS — LISTAR
# ============================================================
def usuarios(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _es_admin(request):
        return _redirigir_dashboard(request)

    try:
        busqueda = request.GET.get('busqueda', '')
        rol      = request.GET.get('rol', '')
        estado   = request.GET.get('estado', '')

        lista = Usuario.objects.all().order_by('nombre')
        if busqueda:
            lista = lista.filter(nombre__icontains=busqueda)
        if rol:
            lista = lista.filter(rol=rol)
        if estado == 'activo':
            lista = lista.filter(activo=True)
        elif estado == 'inactivo':
            lista = lista.filter(activo=False)

        context = {
            'usuarios':          lista,
            'usuario_nombre':    request.session.get('usuario_nombre'),
            'usuario_rol':       request.session.get('usuario_rol'),
            'busqueda':          busqueda,
            'rol':               rol,
            'estado':            estado,
            'total_usuarios':    Usuario.objects.count(),
            'usuarios_activos':  Usuario.objects.filter(activo=True).count(),
            'total_conductores': Usuario.objects.filter(rol='CONDUCTOR').count(),
        }
    except Exception as e:
        context = {
            'usuarios':       [],
            'usuario_nombre': request.session.get('usuario_nombre'),
            'usuario_rol':    request.session.get('usuario_rol'),
            'error':          f'Error al cargar usuarios: {str(e)}',
        }

    return render(request, 'usuarios/lista_usuarios.html', context)


# ============================================================
# CRUD USUARIOS — CREAR
# ============================================================
def usuarios_nuevo(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _es_admin(request):
        return _redirigir_dashboard(request)

    if request.method == 'POST':
        try:
            cedula         = request.POST.get('cedula', '').strip()
            tipo_documento = request.POST.get('tipo_documento', '').strip()
            user_name      = request.POST.get('user_name', '').strip()
            password       = request.POST.get('password', '').strip()
            nombre         = request.POST.get('nombre', '').strip()
            email          = request.POST.get('email', '').strip().lower()
            telefono       = request.POST.get('telefono', '').strip()
            rol            = request.POST.get('rol', '').strip()

            if Usuario.objects.filter(cedula=cedula).exists():
                messages.error(request, 'La cédula ya está registrada.')
                return redirect('usuarios_nuevo')
            if Usuario.objects.filter(user_name=user_name).exists():
                messages.error(request, 'El nombre de usuario ya existe.')
                return redirect('usuarios_nuevo')
            if Usuario.objects.filter(email=email).exists():
                messages.error(request, 'El correo ya está registrado.')
                return redirect('usuarios_nuevo')

            password_hash = bcrypt.hashpw(
                password.encode('utf-8'),
                bcrypt.gensalt()
            ).decode('utf-8')

            Usuario.objects.create(
                cedula         = cedula,
                tipo_documento = tipo_documento,
                user_name      = user_name,
                password       = password_hash,
                nombre         = nombre,
                email          = email,
                telefono       = telefono or None,
                rol            = rol,
                activo         = True,
            )
            messages.success(request, f'Usuario {nombre} creado exitosamente.')
            return redirect('usuarios')

        except Exception as e:
            messages.error(request, f'Error al crear usuario: {str(e)}')
            return redirect('usuarios_nuevo')

    return render(request, 'usuarios/nuevo_usuario.html', {
        'usuario_nombre': request.session.get('usuario_nombre'),
        'usuario_rol':    request.session.get('usuario_rol'),
        'fecha_actual':   date.today(),
        'es_nuevo':       True,
    })


# ============================================================
# CRUD USUARIOS — EDITAR
# ============================================================
def usuarios_editar(request, cedula):
    if not _sesion_activa(request):
        return redirect('login')
    if not _es_admin(request):
        return _redirigir_dashboard(request)

    try:
        usuario_edit = Usuario.objects.get(cedula=cedula)
    except Usuario.DoesNotExist:
        messages.error(request, 'Usuario no encontrado.')
        return redirect('usuarios')

    if request.method == 'POST':
        try:
            usuario_edit.nombre   = request.POST.get('nombre', usuario_edit.nombre).strip()
            usuario_edit.email    = request.POST.get('email', usuario_edit.email).strip()
            usuario_edit.telefono = request.POST.get('telefono', '').strip() or None
            usuario_edit.rol      = request.POST.get('rol', usuario_edit.rol)
            usuario_edit.activo   = request.POST.get('activo') == 'true'

            password_nueva = request.POST.get('password', '').strip()
            if password_nueva:
                usuario_edit.password = bcrypt.hashpw(
                    password_nueva.encode('utf-8'),
                    bcrypt.gensalt()
                ).decode('utf-8')

            usuario_edit.save()
            messages.success(request, f'Usuario {usuario_edit.nombre} actualizado.')
            return redirect('usuarios')

        except Exception as e:
            messages.error(request, f'Error al actualizar: {str(e)}')

    return render(request, 'usuarios/nuevo_usuario.html', {
        'usuario':        usuario_edit,
        'usuario_nombre': request.session.get('usuario_nombre'),
        'usuario_rol':    request.session.get('usuario_rol'),
        'fecha_actual':   date.today(),
        'es_nuevo':       False,
    })


# ============================================================
# CRUD USUARIOS — ELIMINAR lógico
# ============================================================
def usuarios_eliminar(request, cedula):
    if not _sesion_activa(request):
        return redirect('login')
    if not _es_admin(request):
        return _redirigir_dashboard(request)

    if request.method != 'POST':
        return redirect('usuarios')

    if cedula == request.session.get('usuario_cedula'):
        messages.error(request, 'No puedes desactivar tu propia cuenta.')
        return redirect('usuarios')

    try:
        usuario_del = Usuario.objects.get(cedula=cedula)
        nombre      = usuario_del.nombre
        usuario_del.delete()
        messages.success(request, f'Usuario {nombre} eliminado exitosamente.')
    except Usuario.DoesNotExist:
        messages.error(request, 'Usuario no encontrado.')
    except Exception as e:
        messages.error(request, f'Error al eliminar: {str(e)}')

    return redirect('usuarios')


# ============================================================
# REPORTES — Vista principal (consulta multicriterio)
# ============================================================
def reportes(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'COLEGIO'):
        return _redirigir_dashboard(request)

    try:
        from estudiantes.models import Estudiante
        from pagos.models import Pago

        tipo_reporte = request.GET.get('tipo_reporte', 'USUARIOS')
        estado       = request.GET.get('estado', '')
        rol          = request.GET.get('rol', '')
        busqueda     = request.GET.get('busqueda', '')
        fecha_desde  = request.GET.get('fecha_desde', '')
        fecha_hasta  = request.GET.get('fecha_hasta', '')

        datos = _obtener_reportes(
            tipo_reporte, estado, rol, busqueda, fecha_desde, fecha_hasta
        )

        context = {
            'reportes':          datos,
            'total_registros':   len(datos),
            'tipo_reporte':      tipo_reporte,
            'estado':            estado,
            'rol':               rol,
            'busqueda':          busqueda,
            'fecha_desde':       fecha_desde,
            'fecha_hasta':       fecha_hasta,
            'total_usuarios':    Usuario.objects.count(),
            'usuarios_activos':  Usuario.objects.filter(activo=True).count(),
            'total_estudiantes': Estudiante.objects.count(),
            'total_pagos':       Pago.objects.count(),
            'usuario_nombre':    request.session.get('usuario_nombre'),
            'usuario_rol':       request.session.get('usuario_rol'),
        }

    except Exception as e:
        context = {
            'usuario_nombre':  request.session.get('usuario_nombre'),
            'usuario_rol':     request.session.get('usuario_rol'),
            'error':           f'Error al generar reporte: {str(e)}',
            'reportes':        [],
            'total_registros': 0,
        }

    return render(request, 'reportes/reportes.html', context)


# ============================================================
# AUXILIAR — Consulta multicriterio
# ============================================================
def _obtener_reportes(tipo_reporte, estado, rol, busqueda, fecha_desde, fecha_hasta):
    from estudiantes.models import Estudiante
    from pagos.models import Pago
    from rutas.models import Ruta

    try:
        if tipo_reporte == 'USUARIOS':
            qs = Usuario.objects.all()
            if busqueda:
                qs = qs.filter(nombre__icontains=busqueda)
            if rol:
                qs = qs.filter(rol=rol)
            if estado == 'ACTIVO':
                qs = qs.filter(activo=True)
            elif estado == 'INACTIVO':
                qs = qs.filter(activo=False)
            return list(qs.order_by('nombre'))

        elif tipo_reporte == 'ESTUDIANTES':
            qs = Estudiante.objects.all()
            if busqueda:
                qs = qs.filter(nombre__icontains=busqueda)
            if estado == 'ACTIVO':
                qs = qs.filter(activo=True)
            elif estado == 'INACTIVO':
                qs = qs.filter(activo=False)
            if fecha_desde:
                qs = qs.filter(fecha_registro__gte=fecha_desde)
            if fecha_hasta:
                qs = qs.filter(fecha_registro__lte=fecha_hasta)
            return list(qs.order_by('nombre'))

        elif tipo_reporte == 'PAGOS':
            qs = Pago.objects.all()
            if estado == 'ACTIVO':
                qs = qs.filter(estado='PAGADO')
            elif estado == 'INACTIVO':
                qs = qs.filter(estado='PENDIENTE')
            if fecha_desde:
                qs = qs.filter(fecha_pago__gte=fecha_desde)
            if fecha_hasta:
                qs = qs.filter(fecha_pago__lte=fecha_hasta)
            return list(qs.order_by('-fecha_pago'))

        elif tipo_reporte == 'RUTAS':
            qs = Ruta.objects.all()
            if busqueda:
                qs = qs.filter(nombre__icontains=busqueda)
            if estado == 'ACTIVO':
                qs = qs.filter(activo=True)
            elif estado == 'INACTIVO':
                qs = qs.filter(activo=False)
            return list(qs.order_by('nombre'))

    except Exception as e:
        print(f'Error en _obtener_reportes: {str(e)}')

    return []


# ============================================================
# REPORTES — Descargar PDF
# ============================================================
def reportes_pdf(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'COLEGIO'):
        return _redirigir_dashboard(request)

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib           import colors
        from reportlab.lib.units     import cm
        from reportlab.platypus      import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle

        tipo_reporte = request.GET.get('tipo_reporte', 'USUARIOS')
        estado       = request.GET.get('estado', '')
        rol          = request.GET.get('rol', '')
        busqueda     = request.GET.get('busqueda', '')
        fecha_desde  = request.GET.get('fecha_desde', '')
        fecha_hasta  = request.GET.get('fecha_hasta', '')

        datos  = _obtener_reportes(tipo_reporte, estado, rol, busqueda, fecha_desde, fecha_hasta)
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            leftMargin=1*cm, rightMargin=1*cm,
            topMargin=1.5*cm, bottomMargin=1*cm
        )
        styles = getSampleStyleSheet()
        story  = []

        titulo_style = ParagraphStyle(
            'titulo', parent=styles['Title'],
            fontSize=16, textColor=colors.HexColor('#1e293b')
        )
        story.append(Paragraph(f'RutaEscolar — Reporte de {tipo_reporte.title()}', titulo_style))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(
            f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Total: {len(datos)} registros',
            styles['Normal']
        ))
        story.append(Spacer(1, 0.5*cm))

        if tipo_reporte == 'USUARIOS':
            cabeceras = ['Cédula', 'Nombre', 'Usuario', 'Email', 'Teléfono', 'Rol', 'Estado']
            filas = [[str(d.cedula), d.nombre, d.user_name, d.email,
                      d.telefono or '—', d.rol, 'Activo' if d.activo else 'Inactivo']
                     for d in datos]
        elif tipo_reporte == 'ESTUDIANTES':
            cabeceras = ['Documento', 'Nombre', 'Apellido', 'Grado', 'Institución', 'Estado']
            filas = [[d.documento, d.nombre, d.apellido, d.grado or '—',
                      d.institucion or '—', 'Activo' if d.activo else 'Inactivo']
                     for d in datos]
        elif tipo_reporte == 'PAGOS':
            cabeceras = ['Código', 'Estudiante', 'Monto', 'Mes', 'Fecha Pago', 'Método', 'Estado']
            filas = [[d.codigo, str(d.documento_estudiante_id), f'${d.monto}',
                      f'{d.mes or ""} {d.anio or ""}', str(d.fecha_pago),
                      d.metodo_pago or '—', d.estado]
                     for d in datos]
        elif tipo_reporte == 'RUTAS':
            cabeceras = ['Código', 'Nombre', 'Turno', 'Hora Inicio', 'Hora Fin', 'Capacidad', 'Estado']
            filas = [[d.codigo, d.nombre, d.turno or '—',
                      str(d.hora_inicio or '—'), str(d.hora_fin or '—'),
                      str(d.capacidad_maxima or '—'), 'Activa' if d.activo else 'Inactiva']
                     for d in datos]
        else:
            cabeceras = ['Sin datos']
            filas     = [['—']]

        if not filas:
            filas = [['Sin datos para los filtros seleccionados'] + [''] * (len(cabeceras) - 1)]

        page_w = landscape(A4)[0] - 2*cm
        col_w  = [page_w / len(cabeceras)] * len(cabeceras)
        tabla  = Table([cabeceras] + filas, colWidths=col_w, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#1e293b')),
            ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',       (0, 0), (-1, 0),  9),
            ('ALIGN',          (0, 0), (-1, 0),  'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('FONTSIZE',       (0, 1), (-1, -1), 8),
            ('GRID',           (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',     (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING',  (0, 0), (-1, -1), 4),
        ]))
        story.append(tabla)
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph('RutaEscolar © 2026 | Sistema de Gestión de Transporte Escolar', styles['Normal']))
        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="reporte_{tipo_reporte.lower()}_{date.today()}.pdf"'
        )
        return response

    except Exception as e:
        messages.error(request, f'Error al generar PDF: {str(e)}')
        return redirect('reportes')


# ============================================================
# REPORTES — Descargar Excel
# ============================================================
def reportes_excel(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'COLEGIO'):
        return _redirigir_dashboard(request)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        tipo_reporte = request.GET.get('tipo_reporte', 'USUARIOS')
        estado       = request.GET.get('estado', '')
        rol          = request.GET.get('rol', '')
        busqueda     = request.GET.get('busqueda', '')
        fecha_desde  = request.GET.get('fecha_desde', '')
        fecha_hasta  = request.GET.get('fecha_hasta', '')

        datos = _obtener_reportes(tipo_reporte, estado, rol, busqueda, fecha_desde, fecha_hasta)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tipo_reporte.title()

        header_fill = PatternFill('solid', fgColor='1e293b')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        thin        = Side(style='thin', color='e2e8f0')
        borde       = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill    = PatternFill('solid', fgColor='f8fafc')

        ws.merge_cells('A1:G1')
        ws['A1'] = (
            f'RutaEscolar — Reporte de {tipo_reporte.title()} | '
            f'{datetime.now().strftime("%d/%m/%Y %H:%M")}'
        )
        ws['A1'].font      = Font(bold=True, size=13, color='1e293b')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        ws.merge_cells('A2:G2')
        ws['A2'] = f'Total registros: {len(datos)}'
        ws['A2'].alignment = Alignment(horizontal='center')

        if tipo_reporte == 'USUARIOS':
            cabeceras = ['Cédula', 'Nombre', 'Usuario', 'Email', 'Teléfono', 'Rol', 'Estado']
        elif tipo_reporte == 'ESTUDIANTES':
            cabeceras = ['Documento', 'Nombre', 'Apellido', 'Grado', 'Institución', 'Ruta', 'Estado']
        elif tipo_reporte == 'PAGOS':
            cabeceras = ['Código', 'Estudiante', 'Monto', 'Mes / Año', 'Fecha Pago', 'Método', 'Estado']
        elif tipo_reporte == 'RUTAS':
            cabeceras = ['Código', 'Nombre', 'Turno', 'Hora Inicio', 'Hora Fin', 'Capacidad', 'Estado']
        else:
            cabeceras = []

        for col, cab in enumerate(cabeceras, 1):
            cell           = ws.cell(row=4, column=col, value=cab)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = borde
        ws.row_dimensions[4].height = 22

        for row_idx, d in enumerate(datos, 5):
            if tipo_reporte == 'USUARIOS':
                fila = [d.cedula, d.nombre, d.user_name, d.email,
                        d.telefono or '—', d.rol, 'Activo' if d.activo else 'Inactivo']
            elif tipo_reporte == 'ESTUDIANTES':
                fila = [d.documento, d.nombre, d.apellido, d.grado or '—',
                        d.institucion or '—', str(d.codigo_ruta_id or 'Sin ruta'),
                        'Activo' if d.activo else 'Inactivo']
            elif tipo_reporte == 'PAGOS':
                fila = [d.codigo, str(d.documento_estudiante_id), float(d.monto),
                        f'{d.mes or ""} {d.anio or ""}', str(d.fecha_pago),
                        d.metodo_pago or '—', d.estado]
            elif tipo_reporte == 'RUTAS':
                fila = [d.codigo, d.nombre, d.turno or '—',
                        str(d.hora_inicio or '—'), str(d.hora_fin or '—'),
                        d.capacidad_maxima or '—', 'Activa' if d.activo else 'Inactiva']
            else:
                fila = []

            fill = alt_fill if row_idx % 2 == 0 else None
            for col, valor in enumerate(fila, 1):
                cell           = ws.cell(row=row_idx, column=col, value=valor)
                cell.border    = borde
                cell.alignment = Alignment(vertical='center')
                if fill:
                    cell.fill = fill

        for col in ws.columns:
            ancho = 10
            for cell in col:
                try:
                    if cell.value:
                        ancho = max(ancho, len(str(cell.value)))
                except Exception:
                    pass
            try:
                ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 40)
            except AttributeError:
                pass

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="reporte_{tipo_reporte.lower()}_{date.today()}.xlsx"'
        )
        return response

    except Exception as e:
        messages.error(request, f'Error al generar Excel: {str(e)}')
        return redirect('reportes')


# ============================================================
# AUXILIAR — Filtrar usuarios
# ============================================================
def _filtrar_usuarios(request):
    busqueda = request.GET.get('busqueda', '')
    rol      = request.GET.get('rol', '')
    estado   = request.GET.get('estado', '')

    qs = Usuario.objects.all().order_by('nombre')
    if busqueda:
        qs = qs.filter(nombre__icontains=busqueda)
    if rol:
        qs = qs.filter(rol=rol)
    if estado == 'activo':
        qs = qs.filter(activo=True)
    elif estado == 'inactivo':
        qs = qs.filter(activo=False)
    return list(qs)


# ============================================================
# USUARIOS — Descargar PDF
# ============================================================
def usuarios_pdf(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'COLEGIO'):
        return _redirigir_dashboard(request)

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib           import colors
        from reportlab.lib.units     import cm
        from reportlab.platypus      import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle

        datos  = _filtrar_usuarios(request)
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            leftMargin=1*cm, rightMargin=1*cm,
            topMargin=1.5*cm, bottomMargin=1*cm
        )
        styles = getSampleStyleSheet()
        story  = []

        titulo_style = ParagraphStyle(
            'titulo', parent=styles['Title'],
            fontSize=16, textColor=colors.HexColor('#1e293b')
        )
        story.append(Paragraph('RutaEscolar — Reporte de Usuarios', titulo_style))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(
            f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Total: {len(datos)} registros',
            styles['Normal']
        ))
        story.append(Spacer(1, 0.5*cm))

        cabeceras = ['Cédula', 'Nombre', 'Usuario', 'Email', 'Teléfono', 'Rol', 'Estado']
        filas = [
            [str(d.cedula), d.nombre, d.user_name, d.email,
             d.telefono or '—', d.rol, 'Activo' if d.activo else 'Inactivo']
            for d in datos
        ]
        if not filas:
            filas = [['Sin datos'] + [''] * (len(cabeceras) - 1)]

        page_w = landscape(A4)[0] - 2*cm
        col_w  = [page_w / len(cabeceras)] * len(cabeceras)
        tabla  = Table([cabeceras] + filas, colWidths=col_w, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#1e293b')),
            ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',       (0, 0), (-1, 0),  9),
            ('ALIGN',          (0, 0), (-1, 0),  'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('FONTSIZE',       (0, 1), (-1, -1), 8),
            ('GRID',           (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',     (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING',  (0, 0), (-1, -1), 4),
        ]))
        story.append(tabla)
        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="usuarios_{date.today()}.pdf"'
        return response

    except Exception as e:
        messages.error(request, f'Error al generar PDF: {str(e)}')
        return redirect('usuarios')


# ============================================================
# USUARIOS — Descargar Excel
# ============================================================
def usuarios_excel(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'COLEGIO'):
        return _redirigir_dashboard(request)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        datos = _filtrar_usuarios(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Usuarios'

        header_fill = PatternFill('solid', fgColor='1e293b')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        thin        = Side(style='thin', color='e2e8f0')
        borde       = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill    = PatternFill('solid', fgColor='f8fafc')

        ws.merge_cells('A1:G1')
        ws['A1'] = f'RutaEscolar — Usuarios | {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A1'].font      = Font(bold=True, size=13, color='1e293b')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:G2')
        ws['A2'] = f'Total: {len(datos)} registros'
        ws['A2'].alignment = Alignment(horizontal='center')

        cabeceras = ['Cédula', 'Nombre', 'Usuario', 'Email', 'Teléfono', 'Rol', 'Estado']
        for col, cab in enumerate(cabeceras, 1):
            cell           = ws.cell(row=4, column=col, value=cab)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = borde
        ws.row_dimensions[4].height = 22

        for row_idx, d in enumerate(datos, 5):
            fila = [d.cedula, d.nombre, d.user_name, d.email,
                    d.telefono or '—', d.rol, 'Activo' if d.activo else 'Inactivo']
            fill = alt_fill if row_idx % 2 == 0 else None
            for col, valor in enumerate(fila, 1):
                cell           = ws.cell(row=row_idx, column=col, value=valor)
                cell.border    = borde
                cell.alignment = Alignment(vertical='center')
                if fill:
                    cell.fill = fill

        for col in ws.columns:
            ancho = 10
            for cell in col:
                try:
                    if cell.value:
                        ancho = max(ancho, len(str(cell.value)))
                except Exception:
                    pass
            try:
                ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 40)
            except AttributeError:
                pass

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="usuarios_{date.today()}.xlsx"'
        return response

    except Exception as e:
        messages.error(request, f'Error al generar Excel: {str(e)}')
        return redirect('usuarios')


# ============================================================
# MONITORA — Confirmar lista
# ============================================================
def confirmar_lista(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'MONITORA'):
        return _redirigir_dashboard(request)

    if request.method == 'POST':
        try:
            from rutas.models import Recorrido
            from monitoras.models import Monitora

            cedula   = request.session.get('usuario_cedula')
            monitora = Monitora.objects.filter(usuario__cedula=cedula).first()

            recorrido = None
            if monitora and monitora.ruta_asignada:
                recorrido = Recorrido.objects.filter(
                    ruta  = monitora.ruta_asignada,
                    fecha = date.today(),
                ).exclude(estado='CANCELADO').first()

            if recorrido:
                recorrido.lista_confirmada = True
                recorrido.save()

            request.session['lista_confirmada'] = True
            messages.success(request, '✅ Lista confirmada. El conductor ya puede iniciar el recorrido.')

        except Exception as e:
            messages.error(request, f'Error al confirmar lista: {str(e)}')

    return redirect('dashboard_monitora')


# ============================================================
# MONITORA — Registrar novedad por estudiante
# ============================================================
def registrar_novedad(request, documento):
    if not _sesion_activa(request):
        return redirect('login')

    if request.method == 'POST':
        try:
            from estudiantes.models import Estudiante
            estado     = request.POST.get('estado', 'PRESENTE')
            estudiante = Estudiante.objects.get(documento=documento)
            messages.success(request, f'{estudiante.nombre} marcado como {estado}.')
        except Exception as e:
            messages.error(request, f'Error al registrar novedad: {str(e)}')

    return redirect('dashboard_monitora')


# ============================================================
# MONITORA — Registrar novedad general
# ============================================================
def registrar_novedad_general(request):
    if not _sesion_activa(request):
        return redirect('login')

    if request.method == 'POST':
        try:
            from .models import Novedad
            from monitoras.models import Monitora
            from rutas.models import Ruta

            cedula       = request.session.get('usuario_cedula')
            rol          = request.session.get('usuario_rol')
            tipo_novedad = request.POST.get('tipo_novedad', '').strip()
            descripcion  = request.POST.get('descripcion', '').strip()

            if not tipo_novedad or not descripcion:
                messages.error(request, 'Tipo y descripción son obligatorios.')
                return redirect('dashboard_monitora')

            conductor_obj = None
            ruta_obj      = None

            if rol == 'MONITORA':
                monitora = Monitora.objects.filter(usuario__cedula=cedula).first()
                if monitora and monitora.ruta_asignada:
                    ruta_obj = monitora.ruta_asignada
                    if ruta_obj.conductor_cedula:
                        conductor_obj = ruta_obj.conductor_cedula

            novedad = Novedad.objects.create(
                tipo        = tipo_novedad,
                descripcion = descripcion,
                conductor   = conductor_obj,
                ruta        = ruta_obj,
            )

            _notificar_novedad(novedad, request)

            messages.success(request, f'✅ Novedad "{novedad.get_tipo_display()}" registrada y notificada.')

        except Exception as e:
            messages.error(request, f'Error al registrar novedad: {str(e)}')

    return redirect('dashboard_monitora')


# ============================================================
# MONITORA — Solicitar autorización datos médicos
# ============================================================
def solicitar_autorizacion_medica(request, documento):
    if not _sesion_activa(request):
        return redirect('login')

    if request.method == 'POST':
        try:
            from estudiantes.models import Estudiante
            estudiante = Estudiante.objects.get(documento=documento)
            messages.success(request, f'Solicitud enviada al padre de {estudiante.nombre}.')
        except Exception as e:
            messages.error(request, f'Error al enviar solicitud: {str(e)}')

    return redirect('dashboard_monitora')


# ============================================================
# NOTIFICACIONES
#
# NOTA: estas funciones son candidatas naturales a vivir en
# notificaciones/views.py. Se mantienen aquí para no romper urls.py.
# Si ya las migraste, elimina esta sección para evitar duplicados.
# ============================================================
def notificaciones(request):
    if not _sesion_activa(request):
        return redirect('login')

    try:
        from notificaciones.models import Notificacion

        cedula = request.session.get('usuario_cedula')
        filtro = request.GET.get('tipo', '')
        rol    = request.session.get('usuario_rol')

        if rol == 'PADRE':
            notifs = Notificacion.objects.filter(destinatario__cedula=cedula)
        else:
            notifs = Notificacion.objects.all()

        if filtro:
            notifs = notifs.filter(tipo=filtro)

        no_leidas = notifs.exclude(estado='LEIDA').count()

        context = {
            'notificaciones':       notifs,
            'total_notificaciones': notifs.count(),
            'no_leidas':            no_leidas,
            'filtro':               filtro,
            'usuario_nombre':       request.session.get('usuario_nombre'),
            'usuario_rol':          rol,
        }
    except Exception as e:
        context = {
            'notificaciones':       [],
            'total_notificaciones': 0,
            'no_leidas':            0,
            'filtro':               '',
            'usuario_nombre':       request.session.get('usuario_nombre'),
            'usuario_rol':          request.session.get('usuario_rol'),
            'error':                f'Error: {str(e)}',
        }

    return render(request, 'notificaciones.html', context)


# ============================================================
# NOTIFICACIONES — Marcar como leída
# ============================================================
def marcar_leida(request, notif_id):
    if not _sesion_activa(request):
        return redirect('login')

    # CORRECCIÓN: se agrega validación de método POST (antes aceptaba GET,
    # lo que permitía marcar notificaciones como leídas desde un simple enlace).
    if request.method != 'POST':
        return redirect('notificaciones')

    cedula = request.session['usuario_cedula']

    try:
        from notificaciones.models import Notificacion
        from django.utils import timezone

        # CORRECCIÓN IDOR: filtra por destinatario para que un usuario no pueda
        # marcar como leída la notificación de otro cambiando el ID en la URL.
        if _es_admin_o_colegio(request):
            notif = Notificacion.objects.get(id=notif_id)
        else:
            notif = Notificacion.objects.get(
                id=notif_id,
                destinatario__cedula=cedula
            )

        notif.estado        = 'LEIDA'
        notif.fecha_lectura = timezone.now()
        notif.save()
        messages.success(request, 'Notificación marcada como leída.')

    except Notificacion.DoesNotExist:
        messages.error(request, 'Notificación no encontrada.')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')

    return redirect('notificaciones')


# ============================================================
# NOTIFICACIONES — Marcar todas como leídas
# ============================================================
def marcar_todas_leidas(request):
    if not _sesion_activa(request):
        return redirect('login')
    if request.method == 'POST':
        try:
            from notificaciones.models import Notificacion
            from django.utils import timezone
            cedula = request.session.get('usuario_cedula')
            rol    = request.session.get('usuario_rol')
            if rol == 'PADRE':
                qs = Notificacion.objects.filter(destinatario__cedula=cedula)
            else:
                qs = Notificacion.objects.all()
            qs.exclude(estado='LEIDA').update(
                estado='LEIDA',
                fecha_lectura=timezone.now()
            )
            messages.success(request, 'Todas las notificaciones marcadas como leídas.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    return redirect('notificaciones')


# ============================================================
# NOTIFICACIONES — Responder autorización médica (Padre)
# ============================================================
def responder_autorizacion(request, notif_id):
    if not _sesion_activa(request):
        return redirect('login')
    if request.method == 'POST':
        respuesta = request.POST.get('respuesta', 'RECHAZADA')
        if respuesta == 'APROBADA':
            messages.success(request, 'Acceso a datos médicos autorizado.')
        else:
            messages.info(request, 'Acceso a datos médicos rechazado.')
    return redirect('notificaciones')


# ============================================================
# NOTIFICACIONES — Responder ausencia (Padre)
# ============================================================
def responder_ausencia(request, notif_id):
    if not _sesion_activa(request):
        return redirect('login')
    if request.method == 'POST':
        messages.success(request, 'Ausencia confirmada. La monitora y el conductor han sido notificados.')
    return redirect('notificaciones')


# ============================================================
# REPORTE DE VIAJES
# ============================================================
def reporte_viajes(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'COLEGIO'):
        return _redirigir_dashboard(request)

    from rutas.models import Recorrido

    try:
        desde    = request.GET.get('desde', '')
        hasta    = request.GET.get('hasta', '')
        estado   = request.GET.get('estado', '')
        busqueda = request.GET.get('busqueda', '')

        qs = Recorrido.objects.select_related('ruta', 'conductor').all()

        if desde:
            qs = qs.filter(fecha__gte=desde)
        if hasta:
            qs = qs.filter(fecha__lte=hasta)
        if estado:
            qs = qs.filter(estado=estado)
        if busqueda:
            qs = qs.filter(ruta__nombre__icontains=busqueda)

        reportes = _construir_reporte_viajes(qs)

        context = {
            'reportes':         reportes,
            'total_viajes':     len(reportes),
            'fecha_generacion': datetime.now(),
            'usuario_nombre':   request.session.get('usuario_nombre'),
            'usuario_rol':      request.session.get('usuario_rol'),
            'desde':            desde,
            'hasta':            hasta,
            'estado':           estado,
            'busqueda':         busqueda,
        }
    except Exception as e:
        context = {
            'reportes':         [],
            'total_viajes':     0,
            'fecha_generacion': datetime.now(),
            'usuario_nombre':   request.session.get('usuario_nombre'),
            'usuario_rol':      request.session.get('usuario_rol'),
            'error':            f'Error: {str(e)}',
        }

    return render(request, 'reportes/reporte_viajes.html', context)


def _construir_reporte_viajes(qs):
    """Auxiliar compartido entre vista, PDF y Excel."""
    reportes = []
    for r in qs:
        duracion = '—'
        if r.hora_inicio_real and r.hora_fin_real:
            delta    = r.hora_fin_real - r.hora_inicio_real
            minutos  = int(delta.total_seconds() // 60)
            duracion = f"{minutos // 60}h {minutos % 60}m"
        reportes.append({
            'fecha':                r.fecha,
            'ruta_nombre':          r.ruta.nombre,
            'conductor_nombre':     r.conductor.nombre if r.conductor else '—',
            'hora_inicio':          r.hora_inicio_real.strftime('%H:%M') if r.hora_inicio_real else '—',
            'hora_fin':             r.hora_fin_real.strftime('%H:%M') if r.hora_fin_real else '—',
            'duracion':             duracion,
            'cantidad_estudiantes': r.paradas_recorrido.count(),
            'estado':               r.estado,
        })
    return reportes


# ============================================================
# REPORTE DE VIAJES — PDF
# ============================================================
def reporte_viajes_pdf(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'COLEGIO'):
        return _redirigir_dashboard(request)

    try:
        from rutas.models import Recorrido
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib           import colors
        from reportlab.lib.units     import cm
        from reportlab.platypus      import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle

        desde    = request.GET.get('desde', '')
        hasta    = request.GET.get('hasta', '')
        estado   = request.GET.get('estado', '')
        busqueda = request.GET.get('busqueda', '')

        qs = Recorrido.objects.select_related('ruta', 'conductor').all()
        if desde:    qs = qs.filter(fecha__gte=desde)
        if hasta:    qs = qs.filter(fecha__lte=hasta)
        if estado:   qs = qs.filter(estado=estado)
        if busqueda: qs = qs.filter(ruta__nombre__icontains=busqueda)

        datos  = _construir_reporte_viajes(qs)
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            leftMargin=1*cm, rightMargin=1*cm,
            topMargin=1.5*cm, bottomMargin=1*cm
        )
        styles       = getSampleStyleSheet()
        titulo_style = ParagraphStyle(
            'titulo', parent=styles['Title'],
            fontSize=15, textColor=colors.HexColor('#1e293b')
        )
        story = [
            Paragraph('RutaEscolar — Reporte de Viajes', titulo_style),
            Spacer(1, 0.3*cm),
            Paragraph(
                f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Total: {len(datos)} viajes',
                styles['Normal']
            ),
            Spacer(1, 0.5*cm),
        ]

        cabeceras = ['Fecha', 'Ruta', 'Conductor', 'Inicio', 'Fin', 'Duración', 'Paradas', 'Estado']
        filas = [
            [
                str(d['fecha']),
                d['ruta_nombre'],
                d['conductor_nombre'],
                d['hora_inicio'],
                d['hora_fin'],
                d['duracion'],
                str(d['cantidad_estudiantes']),
                d['estado'],
            ]
            for d in datos
        ]
        if not filas:
            filas = [['Sin datos'] + [''] * 7]

        page_w = landscape(A4)[0] - 2*cm
        col_w  = [page_w * p for p in [0.10, 0.22, 0.18, 0.08, 0.08, 0.10, 0.09, 0.15]]
        tabla  = Table([cabeceras] + filas, colWidths=col_w, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#1e293b')),
            ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',       (0, 0), (-1, 0),  9),
            ('ALIGN',          (0, 0), (-1, 0),  'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('FONTSIZE',       (0, 1), (-1, -1), 8),
            ('GRID',           (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',     (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING',  (0, 0), (-1, -1), 4),
        ]))
        story.append(tabla)
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph('RutaEscolar © 2026', styles['Normal']))
        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="viajes_{date.today()}.pdf"'
        return response

    except Exception as e:
        messages.error(request, f'Error al generar PDF: {str(e)}')
        return redirect('reporte_viajes')


# ============================================================
# REPORTE DE VIAJES — Excel
# ============================================================
def reporte_viajes_excel(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'COLEGIO'):
        return _redirigir_dashboard(request)

    try:
        from rutas.models import Recorrido
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        desde    = request.GET.get('desde', '')
        hasta    = request.GET.get('hasta', '')
        estado   = request.GET.get('estado', '')
        busqueda = request.GET.get('busqueda', '')

        qs = Recorrido.objects.select_related('ruta', 'conductor').all()
        if desde:    qs = qs.filter(fecha__gte=desde)
        if hasta:    qs = qs.filter(fecha__lte=hasta)
        if estado:   qs = qs.filter(estado=estado)
        if busqueda: qs = qs.filter(ruta__nombre__icontains=busqueda)

        datos = _construir_reporte_viajes(qs)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Viajes'

        header_fill = PatternFill('solid', fgColor='1E293B')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        thin        = Side(style='thin', color='E2E8F0')
        borde       = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill    = PatternFill('solid', fgColor='F8FAFC')

        ws.merge_cells('A1:H1')
        ws['A1']           = f'RutaEscolar — Reporte de Viajes | {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A1'].font      = Font(bold=True, size=13, color='1E293B')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        ws.merge_cells('A2:H2')
        ws['A2']           = f'Total: {len(datos)} viajes'
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font      = Font(size=10, color='64748B')

        ws.row_dimensions[3].height = 6

        cabeceras = ['Fecha', 'Ruta', 'Conductor', 'Hora Inicio', 'Hora Fin', 'Duración', 'Paradas', 'Estado']
        for col, cab in enumerate(cabeceras, 1):
            cell           = ws.cell(row=4, column=col, value=cab)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = borde
        ws.row_dimensions[4].height = 22

        for row_idx, d in enumerate(datos, 5):
            fila = [
                str(d['fecha']),
                d['ruta_nombre'],
                d['conductor_nombre'],
                d['hora_inicio'],
                d['hora_fin'],
                d['duracion'],
                d['cantidad_estudiantes'],
                d['estado'],
            ]
            fill = alt_fill if row_idx % 2 == 0 else None
            for col, valor in enumerate(fila, 1):
                cell           = ws.cell(row=row_idx, column=col, value=valor)
                cell.border    = borde
                cell.alignment = Alignment(vertical='center')
                if fill:
                    cell.fill = fill

        anchos = [12, 25, 20, 12, 12, 12, 10, 14]
        letras = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        for letra, ancho in zip(letras, anchos):
            ws.column_dimensions[letra].width = ancho

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="viajes_{date.today()}.xlsx"'
        return response

    except Exception as e:
        messages.error(request, f'Error al generar Excel: {str(e)}')
        return redirect('reporte_viajes')


# ============================================================
# ADMIN — Resetear lista de recorrido
# ============================================================
# CORRECCIÓN CRÍTICA: esta función estaba indentada dentro de
# reporte_viajes_excel, convirtiéndola en una función anidada invisible
# para Django (nunca registrable en urls.py). Se mueve al nivel de módulo.
def resetear_lista_recorrido(request):
    """ADMIN puede resetear la lista confirmada del recorrido del día para pruebas o correcciones."""
    if not _sesion_activa(request):
        return redirect('login')
    if not _es_admin(request):
        return _redirigir_dashboard(request)

    if request.method == 'POST':
        try:
            from rutas.models import Recorrido
            recorrido_id = request.POST.get('recorrido_id', '').strip()

            if recorrido_id:
                rec = Recorrido.objects.get(id=recorrido_id)
            else:
                rec = Recorrido.objects.filter(fecha=date.today()).order_by('-fecha_creacion').first()

            if rec:
                rec.lista_confirmada = False
                rec.save()
                request.session['lista_confirmada'] = False
                messages.success(
                    request,
                    f'✅ Lista del recorrido del {rec.fecha} reseteada. '
                    f'La monitora puede confirmar nuevamente.'
                )
            else:
                messages.warning(request, 'No hay recorrido activo hoy para resetear.')

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    return redirect(request.POST.get('next', 'dashboard_admin'))


# ============================================================
# NOVEDADES — Registrar (Conductor)
# ============================================================
def registrar_novedad_conductor(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'CONDUCTOR'):
        return _redirigir_dashboard(request)

    if request.method == 'POST':
        try:
            from .models import Novedad
            from rutas.models import Ruta
            from estudiantes.models import Estudiante

            cedula        = request.session.get('usuario_cedula')
            conductor_obj = Usuario.objects.get(cedula=cedula)
            tipo          = request.POST.get('tipo', '').strip()
            descripcion   = request.POST.get('descripcion', '').strip()
            ruta_cod      = request.POST.get('ruta_codigo', '').strip()
            est_doc       = request.POST.get('documento_estudiante', '').strip()

            if not tipo or not descripcion:
                messages.error(request, 'Tipo y descripción son obligatorios.')
                return redirect('dashboard_conductor')

            ruta_obj = Ruta.objects.filter(codigo=ruta_cod).first() if ruta_cod else None
            est_obj  = Estudiante.objects.filter(documento=est_doc).first() if est_doc else None

            novedad = Novedad.objects.create(
                tipo        = tipo,
                descripcion = descripcion,
                conductor   = conductor_obj,
                ruta        = ruta_obj,
                estudiante  = est_obj,
            )

            _notificar_novedad(novedad, request)

            messages.success(request, f'✅ Novedad "{novedad.get_tipo_display()}" registrada y notificada.')

        except Exception as e:
            messages.error(request, f'Error al registrar novedad: {str(e)}')

    return redirect('dashboard_conductor')


def _notificar_novedad(novedad, request):
    """Envía notificación en la app a ADMIN y COLEGIO sobre la novedad."""
    try:
        from notificaciones.models import Notificacion

        destinatarios = Usuario.objects.filter(
            rol__in=('ADMIN', 'COLEGIO'), activo=True
        )

        for dest in destinatarios:
            Notificacion.objects.create(
                destinatario = dest,
                tipo         = 'AVISO_GENERAL',
                canal        = 'APP',
                estado       = 'ENVIADA',
                mensaje      = (
                    f'🚨 Novedad: {novedad.get_tipo_display()} — '
                    f'{novedad.descripcion[:100]}'
                    f'{(" — Ruta: " + novedad.ruta.nombre) if novedad.ruta else ""}'
                ),
            )
    except Exception as e:
        print(f'⚠️ Error notificando novedad: {e}')


# ============================================================
# AUSENCIAS — Reportar (Padre)
# ============================================================
def reportar_ausencia(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'PADRE'):
        return _redirigir_dashboard(request)

    if request.method == 'POST':
        try:
            from .models import AusenciaEstudiante
            from estudiantes.models import Estudiante
            from notificaciones.models import Notificacion

            cedula    = request.session.get('usuario_cedula')
            padre_obj = Usuario.objects.get(cedula=cedula)
            est_doc   = request.POST.get('documento_estudiante', '').strip()
            fecha_aus = request.POST.get('fecha', '').strip()
            razon     = request.POST.get('razon', '').strip()

            if not est_doc or not fecha_aus:
                messages.error(request, 'Estudiante y fecha son obligatorios.')
                return redirect('dashboard_padre')

            estudiante = Estudiante.objects.get(documento=est_doc, cedula_padre=cedula)

            if AusenciaEstudiante.objects.filter(estudiante=estudiante, fecha=fecha_aus).exists():
                messages.warning(
                    request,
                    f'Ya existe una ausencia registrada para {estudiante.nombre} en esa fecha.'
                )
                return redirect('dashboard_padre')

            AusenciaEstudiante.objects.create(
                estudiante = estudiante,
                padre      = padre_obj,
                fecha      = fecha_aus,
                razon      = razon or None,
                estado     = 'CONFIRMADA',
            )

            _notificar_ausencia(estudiante, fecha_aus, razon)

            messages.success(
                request,
                f'✅ Ausencia de {estudiante.nombre} reportada para {fecha_aus}. '
                f'Se notificó a la monitora y conductor.'
            )

        except Estudiante.DoesNotExist:
            messages.error(request, 'Estudiante no encontrado o no asociado a tu cuenta.')
        except Exception as e:
            messages.error(request, f'Error al reportar ausencia: {str(e)}')

    return redirect('dashboard_padre')


def _notificar_ausencia(estudiante, fecha, razon):
    """Notifica a monitora y conductor de la ruta del estudiante."""
    try:
        from notificaciones.models import Notificacion
        from monitoras.models import Monitora

        if not estudiante.codigo_ruta:
            return

        ruta = estudiante.codigo_ruta
        mensaje = (
            f'📋 Ausencia: {estudiante.nombre} {estudiante.apellido} '
            f'no asistirá el {fecha}.'
            f'{(" Motivo: " + razon) if razon else ""}'
        )

        monitoras = Monitora.objects.filter(
            ruta_asignada=ruta,
            usuario__activo=True
        ).select_related('usuario')

        for m in monitoras:
            Notificacion.objects.create(
                destinatario = m.usuario,
                tipo         = 'AVISO_GENERAL',
                canal        = 'APP',
                estado       = 'ENVIADA',
                mensaje      = mensaje,
            )

        if ruta.conductor_cedula:
            Notificacion.objects.create(
                destinatario = ruta.conductor_cedula,
                tipo         = 'AVISO_GENERAL',
                canal        = 'APP',
                estado       = 'ENVIADA',
                mensaje      = mensaje,
            )

    except Exception as e:
        print(f'⚠️ Error notificando ausencia: {e}')


# ============================================================
# NOVEDADES — Lista (ADMIN/COLEGIO)
# ============================================================
def lista_novedades(request):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'COLEGIO'):
        return _redirigir_dashboard(request)

    try:
        from .models import Novedad

        tipo   = request.GET.get('tipo', '')
        estado = request.GET.get('estado', '')
        desde  = request.GET.get('desde', '')
        hasta  = request.GET.get('hasta', '')

        qs = Novedad.objects.select_related('conductor', 'ruta', 'estudiante').all()

        if tipo:
            qs = qs.filter(tipo=tipo)
        if estado:
            qs = qs.filter(estado=estado)
        if desde:
            qs = qs.filter(fecha_hora__date__gte=desde)
        if hasta:
            qs = qs.filter(fecha_hora__date__lte=hasta)

        context = {
            'novedades':      qs,
            'total':          qs.count(),
            'tipos':          Novedad.TIPO,
            'estados':        Novedad.ESTADO,
            'tipo':           tipo,
            'estado':         estado,
            'desde':          desde,
            'hasta':          hasta,
            'usuario_nombre': request.session.get('usuario_nombre'),
            'usuario_rol':    request.session.get('usuario_rol'),
            'fecha_actual':   date.today(),
        }
    except Exception as e:
        context = {
            'novedades':      [],
            'total':          0,
            'tipos':          [],
            'estados':        [],
            'usuario_nombre': request.session.get('usuario_nombre'),
            'usuario_rol':    request.session.get('usuario_rol'),
            'error':          f'Error: {str(e)}',
        }

    return render(request, 'novedades/lista_novedades.html', context)
def resolver_novedad(request, novedad_id):
    if not _sesion_activa(request):
        return redirect('login')
    if not _rol_en(request, 'COLEGIO'):
        return _redirigir_dashboard(request)
    if request.method == 'POST':
        try:
            from .models import Novedad
            novedad        = Novedad.objects.get(id=novedad_id)
            novedad.estado = 'RESUELTA'
            novedad.save()
            messages.success(request, 'Novedad marcada como resuelta.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    return redirect('novedades')
